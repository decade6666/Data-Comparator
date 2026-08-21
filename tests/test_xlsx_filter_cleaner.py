import zipfile
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo

from src.backend.infrastructure import xlsx_filter_cleaner as cleaner
from src.backend.infrastructure.xlsx_filter_cleaner import (
    CleanupOptions,
    NotAnOoxmlPackageError,
    remove_filters,
    strip_autofilter,
    strip_filter_database,
    unhide_rows,
)


@pytest.fixture
def filtered_workbook(tmp_path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "S1"
    sheet.append(["id", "value", "note"])
    for row in range(1, 6):
        sheet.append([row, row * 10, f"row-{row}"])
    sheet.auto_filter.ref = "A1:C6"
    sheet.row_dimensions[4].hidden = True
    sheet.column_dimensions["B"].hidden = True

    table_sheet = workbook.create_sheet("S2")
    table_sheet.append(["id", "value"])
    for row in range(1, 4):
        table_sheet.append([row, row * 10])
    table = Table(displayName="T1", ref="A1:B4")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False
    )
    table_sheet.add_table(table)

    workbook.defined_names.add(
        DefinedName(
            "_xlnm._FilterDatabase",
            attr_text="'S1'!$A$1:$C$6",
            localSheetId=0,
        )
    )
    path = tmp_path / "filtered.xlsx"
    workbook.save(path)
    return path


@pytest.mark.parametrize(
    "xml",
    [
        b'<autoFilter ref="A1:B2"/>',
        b'<autoFilter ref="A1:B2"><filterColumn colId="0"/></autoFilter>',
        b'<x:autoFilter ref="A1:B2"/>',
        b'<autoFilter ref="A1:B2" note="a > b"/>',
        b'<autoFilter ref="A1:B2"></autoFilter >',
    ],
)
def test_strip_autofilter_removes_supported_shapes(xml: bytes) -> None:
    cleaned, count = strip_autofilter(b"<worksheet>" + xml + b"</worksheet>")

    assert count == 1
    assert b"autoFilter" not in cleaned


def test_strip_autofilter_preserves_namespace_bytes() -> None:
    xml = (
        b'<?xml version="1.0"?>'
        b'<worksheet xmlns="http://schemas.openxmlformats.org/'
        b'spreadsheetml/2006/main" '
        b'xmlns:mc="http://schemas.openxmlformats.org/'
        b'markup-compatibility/2006" '
        b'xmlns:x14ac="http://schemas.microsoft.com/office/'
        b'spreadsheetml/2009/9/ac" '
        b'xmlns:xr="http://schemas.microsoft.com/office/'
        b'spreadsheetml/2014/revision" '
        b'xmlns:xr2="http://schemas.microsoft.com/office/'
        b'spreadsheetml/2015/revision2" '
        b'xmlns:xr3="http://schemas.microsoft.com/office/'
        b'spreadsheetml/2016/revision3" '
        b'mc:Ignorable="x14ac xr xr2 xr3">'
        b'<autoFilter ref="A1:B2"/></worksheet>'
    )

    cleaned, count = strip_autofilter(xml)
    root_start = xml[: xml.index(b">", xml.index(b"<worksheet")) + 1]

    assert count == 1
    assert cleaned.startswith(root_start)
    assert b"ns0:" not in cleaned
    assert b'xmlns:xr2="' in cleaned
    assert b'mc:Ignorable="x14ac xr xr2 xr3"' in cleaned


def test_strip_autofilter_ignores_comments_and_cdata() -> None:
    xml = (
        b'<worksheet><!-- <autoFilter ref="A1:B2"/> -->'
        b'<![CDATA[<autoFilter ref="C1:D2"/>]]></worksheet>'
    )

    cleaned, count = strip_autofilter(xml)

    assert count == 0
    assert cleaned == xml


def test_unhide_rows_does_not_parse_attribute_values() -> None:
    xml = b'<row note=\' hidden="1"\' hidden="true"/>'

    cleaned, count = unhide_rows(xml)

    assert count == 1
    assert cleaned == b"<row note=' hidden=\"1\"'/>"


def test_strip_filter_database_does_not_parse_attribute_values() -> None:
    kept_start = b"<definedName note=' name=\"_xlnm._FilterDatabase\" '>"
    kept = kept_start + b"keep</definedName>"
    target = b'<definedName name="_xlnm._FilterDatabase">remove</definedName>'
    xml = kept + target

    cleaned, count = strip_filter_database(xml)

    assert count == 1
    assert cleaned == kept


@pytest.mark.parametrize(
    "xml, expected_count",
    [
        (b'<row r="1" hidden="1"/>', 1),
        (b'<x:row r="2" hidden="true" ht="15">', 1),
        (b'<row r="3" hidden="0"/>', 0),
        (b'<row r="4"/>', 0),
        (b'<rowBreaks><brk hidden="1"/></rowBreaks>', 0),
        (b'<c><is><t>hidden="1"</t></is></c>', 0),
    ],
)
def test_unhide_rows_only_removes_hidden_row_attributes(
    xml: bytes, expected_count: int
) -> None:
    cleaned, count = unhide_rows(xml)

    assert count == expected_count
    if expected_count:
        assert b"hidden" not in cleaned
    else:
        assert cleaned == xml


def test_strip_filter_database_removes_only_target_defined_name() -> None:
    xml = (
        b'<definedNames><definedName name="Keep">Sheet!$A$1</definedName>'
        b'<definedName localSheetId="0" name="_xlnm._FilterDatabase">'
        b"'S1'!$A$1:$B$2</definedName></definedNames>"
    )

    cleaned, count = strip_filter_database(xml)

    assert count == 1
    assert b'name="Keep"' in cleaned
    assert b"_FilterDatabase" not in cleaned


def test_remove_filters_roundtrip_preserves_workbook_data(
    filtered_workbook: Path,
) -> None:
    before = _read_rows(filtered_workbook)
    result = remove_filters(str(filtered_workbook))

    workbook = load_workbook(filtered_workbook)
    try:
        assert workbook.sheetnames == ["S1", "S2"]
        assert workbook["S1"].auto_filter.ref is None
        assert workbook["S1"].row_dimensions[4].hidden is False
        assert workbook["S1"].column_dimensions["B"].hidden is True
        assert "T1" in workbook["S2"].tables
    finally:
        workbook.close()

    assert result.rewritten is True
    assert result.parts_modified >= 2
    assert result.sheet_autofilters_removed == 1
    assert result.table_autofilters_removed == 1
    assert result.hidden_rows_restored == 1
    assert _read_rows(filtered_workbook) == before

    with zipfile.ZipFile(filtered_workbook) as package:
        assert b"autoFilter" not in package.read("xl/worksheets/sheet1.xml")
        assert b"autoFilter" not in package.read("xl/tables/table1.xml")


def test_remove_filters_preserves_entry_order_and_compression(
    filtered_workbook: Path,
) -> None:
    with zipfile.ZipFile(filtered_workbook, mode="a") as source:
        source.comment = b"package comment"
    with zipfile.ZipFile(filtered_workbook) as source:
        source_names = source.namelist()
        source_compression = [info.compress_type for info in source.infolist()]
        source_metadata = [
            (
                info.date_time,
                info.external_attr,
                info.internal_attr,
                info.create_system,
                info.comment,
                info.extra,
            )
            for info in source.infolist()
        ]

    remove_filters(str(filtered_workbook))

    with zipfile.ZipFile(filtered_workbook) as cleaned:
        assert cleaned.namelist() == source_names
        assert [info.compress_type for info in cleaned.infolist()] == (
            source_compression
        )
        assert [
            (
                info.date_time,
                info.external_attr,
                info.internal_attr,
                info.create_system,
                info.comment,
                info.extra,
            )
            for info in cleaned.infolist()
        ] == source_metadata
        assert cleaned.comment == b"package comment"


def test_remove_filters_only_rewrites_dirty_duplicate_entry(
    tmp_path: Path,
) -> None:
    path = tmp_path / "duplicate.xlsx"
    clean_sheet = b"<worksheet><sheetData/></worksheet>"
    filtered_sheet = b'<worksheet><autoFilter ref="A1:B2"/></worksheet>'
    with zipfile.ZipFile(path, "w") as package:
        package.writestr("xl/worksheets/sheet1.xml", clean_sheet)
        with pytest.warns(UserWarning, match="Duplicate name"):
            package.writestr("xl/worksheets/sheet1.xml", filtered_sheet)

    with pytest.warns(UserWarning, match="Duplicate name"):
        result = remove_filters(str(path))

    with zipfile.ZipFile(path) as package:
        entries = [package.read(info) for info in package.infolist()]
    assert entries == [clean_sheet, b"<worksheet></worksheet>"]
    assert result.parts_modified == 1


def test_clean_workbook_is_not_rewritten(tmp_path: Path) -> None:
    workbook = Workbook()
    workbook.active.append(["id", "value"])
    path = tmp_path / "clean.xlsx"
    workbook.save(path)
    before = path.read_bytes()
    before_stat = path.stat()

    result = remove_filters(str(path))

    after_stat = path.stat()
    assert result.rewritten is False
    assert path.read_bytes() == before
    assert after_stat.st_mtime_ns == before_stat.st_mtime_ns
    assert after_stat.st_size == before_stat.st_size


def test_comment_only_workbook_is_not_rewritten(tmp_path: Path) -> None:
    path = tmp_path / "comment-only.xlsx"
    with zipfile.ZipFile(path, "w") as package:
        package.writestr(
            "xl/worksheets/sheet1.xml",
            b'<worksheet><!-- <autoFilter ref="A1:B2"/> --></worksheet>',
        )
    before = path.read_bytes()

    result = remove_filters(str(path))

    assert result.rewritten is False
    assert path.read_bytes() == before


def test_non_zip_input_raises_clear_error(tmp_path: Path) -> None:
    path = tmp_path / "legacy.xls"
    path.write_bytes(b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1" + b"legacy")

    with pytest.raises(NotAnOoxmlPackageError, match="OOXML"):
        remove_filters(str(path))


@pytest.mark.parametrize(
    "attrs",
    [b" ", b"=", b"foo bar", b"foo=bar", b"foo= ", b'foo="unterminated'],
)
def test_attribute_scanner_ignores_malformed_fragments(attrs: bytes) -> None:
    assert list(cleaner._iter_attributes(attrs)) == []


def test_rewrite_row_tag_returns_unmatched_input() -> None:
    tag = b"<not-a-row>"

    assert cleaner._rewrite_row_tag(tag) == (tag, 0)


def test_strip_filter_database_returns_unchanged_without_matches() -> None:
    xml = b"<definedNames/>"

    assert strip_filter_database(xml) == (xml, 0)


def test_read_entry_converts_package_read_errors() -> None:
    class BrokenArchive:
        def open(self, _info):
            raise zipfile.BadZipFile("broken")

    class Info:
        filename = "xl/worksheets/sheet1.xml"

    with pytest.raises(NotAnOoxmlPackageError, match="无法读取 OOXML 部件"):
        cleaner._read_entry(BrokenArchive(), Info())


def test_new_temp_path_returns_open_file_in_same_directory(
    tmp_path: Path,
) -> None:
    source_path = tmp_path / "source.xlsx"
    descriptor, temporary_path = cleaner._new_temp_path(str(source_path))
    try:
        assert Path(temporary_path).parent == tmp_path
        assert Path(temporary_path).name.startswith(".source.xlsx.rewrite-")
    finally:
        cleaner.os.close(descriptor)
        cleaner.os.remove(temporary_path)


def test_cleanup_temp_path_logs_cleanup_failure(monkeypatch) -> None:
    messages = []

    def fail_remove(_path):
        raise OSError("remove failed")

    monkeypatch.setattr(cleaner.os, "remove", fail_remove)

    cleaner._cleanup_temp_path("temporary", messages.append)

    assert any("清理 OOXML 重写临时文件失败" in message for message in messages)


def test_rewrite_package_read_error_preserves_source_and_temp(
    filtered_workbook: Path, monkeypatch
) -> None:
    before = filtered_workbook.read_bytes()

    def fail_rewrite(*_args, **_kwargs):
        raise zipfile.BadZipFile("broken during rewrite")

    monkeypatch.setattr(cleaner, "_rewrite_package", fail_rewrite)

    with pytest.raises(NotAnOoxmlPackageError, match="无法重写 OOXML 文件"):
        remove_filters(str(filtered_workbook))

    assert filtered_workbook.read_bytes() == before
    pattern = "filtered.xlsx.rewrite-*.tmp"
    assert list(filtered_workbook.parent.glob(pattern)) == []


def test_missing_worksheets_part_is_a_noop(tmp_path: Path) -> None:
    path = tmp_path / "partial.xlsx"
    with zipfile.ZipFile(path, "w") as package:
        package.writestr("hello.txt", "not an OOXML worksheet")
    before = path.read_bytes()

    result = remove_filters(str(path))

    assert result.rewritten is False
    assert result.parts_scanned == 0
    assert path.read_bytes() == before


def test_interrupted_cleanup_preserves_source_and_temp_files(
    filtered_workbook: Path, monkeypatch
) -> None:
    before = filtered_workbook.read_bytes()
    calls = 0

    def stop_on_rewrite(_log_func, _stop_flag) -> None:
        nonlocal calls
        calls += 1
        if calls > 1:
            raise InterruptedError("用户停止了操作")

    monkeypatch.setattr(cleaner, "check_stop_frequently", stop_on_rewrite)

    with pytest.raises(InterruptedError, match="用户停止了操作"):
        remove_filters(str(filtered_workbook))

    assert filtered_workbook.read_bytes() == before
    pattern = "filtered.xlsx.rewrite-*.tmp"
    temporary_files = list(filtered_workbook.parent.glob(pattern))
    assert temporary_files == []


def test_write_failure_preserves_source_and_cleans_temp(
    filtered_workbook: Path, monkeypatch
) -> None:
    before = filtered_workbook.read_bytes()

    def fail_replace(_source, _target) -> None:
        raise OSError("替换失败")

    monkeypatch.setattr(cleaner.os, "replace", fail_replace)

    with pytest.raises(OSError, match="替换失败"):
        remove_filters(str(filtered_workbook))

    assert filtered_workbook.read_bytes() == before
    pattern = "filtered.xlsx.rewrite-*.tmp"
    temporary_files = list(filtered_workbook.parent.glob(pattern))
    assert temporary_files == []


def test_filter_database_is_preserved_by_default(
    filtered_workbook: Path,
) -> None:
    remove_filters(str(filtered_workbook))

    with zipfile.ZipFile(filtered_workbook) as package:
        assert b"_FilterDatabase" in package.read("xl/workbook.xml")


def test_filter_database_can_be_removed_explicitly(
    filtered_workbook: Path,
) -> None:
    options = CleanupOptions(remove_filter_database=True)

    result = remove_filters(str(filtered_workbook), options=options)

    with zipfile.ZipFile(filtered_workbook) as package:
        assert b"_FilterDatabase" not in package.read("xl/workbook.xml")
    workbook = load_workbook(filtered_workbook)
    try:
        assert workbook.sheetnames == ["S1", "S2"]
    finally:
        workbook.close()
    assert result.defined_names_removed == 2


def _read_rows(path: Path) -> list:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return {
            sheet.title: list(sheet.iter_rows(values_only=True))
            for sheet in workbook.worksheets
        }
    finally:
        workbook.close()
