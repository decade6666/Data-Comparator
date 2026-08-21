from pathlib import Path

import pytest
from openpyxl import Workbook

from src.backend.infrastructure import file_runtime


def test_check_and_remove_file_protection_processes_xlsx_copy(
    tmp_path: Path,
) -> None:
    source = _write_workbook(tmp_path / "source.xlsx")
    original_bytes = source.read_bytes()
    messages = []

    result = file_runtime.check_and_remove_file_protection(
        str(source), [], messages.append, work_dir=str(tmp_path)
    )

    assert result[0:2] == (False, False)
    assert result[2] != str(source)
    assert result[3] is not None
    assert result[3].rewritten is False
    assert source.read_bytes() == original_bytes
    assert any("未发现自动筛选器" in message for message in messages)


def test_check_and_remove_file_protection_logs_rewritten_cleanup(
    tmp_path: Path,
) -> None:
    source = _write_workbook(tmp_path / "filtered.xlsx")
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet"
    sheet.append(["id", "value"])
    sheet.append([1, 10])
    sheet.auto_filter.ref = "A1:B2"
    sheet.row_dimensions[2].hidden = True
    workbook.save(source)
    workbook.close()
    original_bytes = source.read_bytes()
    messages = []

    result = file_runtime.check_and_remove_file_protection(
        str(source), [], messages.append, work_dir=str(tmp_path)
    )

    assert result[3] is not None and result[3].rewritten is True
    assert source.read_bytes() == original_bytes
    assert any("已清除筛选器" in message for message in messages)


def test_check_and_remove_file_protection_propagates_copy_stop(
    monkeypatch, tmp_path: Path
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")

    def stop_copy(*_args, **_kwargs):
        raise InterruptedError("用户停止了操作")

    monkeypatch.setattr(file_runtime, "check_stop_frequently", stop_copy)

    with pytest.raises(InterruptedError, match="用户停止了操作"):
        file_runtime.check_and_remove_file_protection(
            str(source), [], lambda _message: None, work_dir=str(tmp_path)
        )


def test_check_and_remove_file_protection_requires_existing_file(
    tmp_path: Path,
) -> None:
    with pytest.raises(FileNotFoundError, match="文件不存在"):
        file_runtime.check_and_remove_file_protection(
            str(tmp_path / "missing.xlsx"), [], lambda _message: None
        )


def test_check_and_remove_file_protection_propagates_copy_error(
    monkeypatch, tmp_path: Path
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    messages = []

    def fail_copy(*_args, **_kwargs):
        raise OSError("copy failed")

    monkeypatch.setattr(file_runtime.shutil, "copy2", fail_copy)

    with pytest.raises(OSError, match="copy failed"):
        file_runtime.check_and_remove_file_protection(
            str(source), [], messages.append, work_dir=str(tmp_path)
        )
    assert any("创建副本失败" in message for message in messages)


def test_check_and_remove_file_protection_cleans_copy_on_cleanup_error(
    monkeypatch, tmp_path: Path
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")

    def fail_cleanup(*_args, **_kwargs):
        raise OSError("cleanup failed")

    monkeypatch.setattr(file_runtime, "remove_filters", fail_cleanup)

    with pytest.raises(OSError, match="cleanup failed"):
        file_runtime.check_and_remove_file_protection(
            str(source), [], None, work_dir=str(tmp_path)
        )

    assert list(tmp_path.glob("source_nofilter_*.xlsx")) == []


def test_validate_excel_file_reports_missing_and_empty(tmp_path: Path) -> None:
    messages = []
    missing = tmp_path / "missing.xlsx"
    empty = tmp_path / "empty.xlsx"
    empty.write_bytes(b"")

    assert file_runtime.validate_excel_file(str(missing), messages.append) == (
        False,
        "文件不存在: " + str(missing),
    )
    assert file_runtime.validate_excel_file(str(empty), messages.append) == (
        False,
        "文件为空: " + str(empty),
    )


def test_validate_excel_file_returns_success_from_engine(
    monkeypatch, tmp_path: Path
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    monkeypatch.setattr(
        file_runtime.pd,
        "read_excel",
        lambda *args, **kwargs: object(),
    )

    result = file_runtime.validate_excel_file(
        str(source),
        lambda _message: None,
    )

    assert result == (
        True,
        None,
    )


def test_validate_excel_file_reports_engine_failures(
    monkeypatch, tmp_path: Path
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    messages = []

    def fail_engine(*_args, **_kwargs):
        raise ValueError("invalid XML")

    monkeypatch.setattr(file_runtime.pd, "read_excel", fail_engine)

    valid, error = file_runtime.validate_excel_file(
        str(source),
        messages.append,
    )

    assert valid is False
    assert error is not None and "openpyxl" in error
    assert any("XML格式错误" in message for message in messages)
    assert any("所有验证引擎都失败" in message for message in messages)


def test_validate_excel_file_outer_error(monkeypatch, tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    messages = []

    def fail_exists(_path):
        raise RuntimeError("stat failed")

    monkeypatch.setattr(file_runtime.os.path, "exists", fail_exists)

    valid, error = file_runtime.validate_excel_file(
        str(source),
        messages.append,
    )

    assert valid is False
    assert error == "文件验证过程出错: stat failed"
    assert messages == [error]


def test_get_sheet_names_reads_workbook(tmp_path: Path) -> None:
    source = _write_workbook(tmp_path / "source.xlsx")

    sheet_names = file_runtime.get_sheet_names(
        str(source),
        lambda _message: None,
    )

    assert sheet_names == ["Sheet"]


def test_get_sheet_names_pandas_fallback(monkeypatch, tmp_path: Path) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    messages = []
    closed = []

    class FakeExcelFile:
        sheetnames = ["Fallback"]

        def close(self):
            closed.append(True)

    monkeypatch.setattr(
        file_runtime,
        "load_workbook",
        lambda *_args, **_kwargs: _raise_value_error(),
    )
    monkeypatch.setattr(
        file_runtime.pd, "ExcelFile", lambda *_args, **_kwargs: FakeExcelFile()
    )

    sheet_names = file_runtime.get_sheet_names(
        str(source),
        messages.append,
    )

    assert sheet_names == ["Fallback"]
    assert closed == [True]
    assert any("无法获取文件" in message for message in messages)


def test_get_sheet_names_returns_empty_when_fallback_fails(
    monkeypatch, tmp_path: Path
) -> None:
    source = tmp_path / "source.xlsx"
    source.write_bytes(b"source")
    messages = []

    monkeypatch.setattr(
        file_runtime,
        "load_workbook",
        lambda *_args, **_kwargs: _raise_value_error(),
    )
    monkeypatch.setattr(
        file_runtime.pd,
        "ExcelFile",
        lambda *_args, **_kwargs: _raise_value_error(),
    )

    sheet_names = file_runtime.get_sheet_names(
        str(source),
        messages.append,
    )

    assert sheet_names == []
    assert any("pandas回退也失败" in message for message in messages)


def test_cleanup_nofilter_files_removes_matching_files(tmp_path: Path) -> None:
    candidate = tmp_path / "source_nofilter_1.xlsx"
    candidate.write_bytes(b"temporary")
    legacy_candidate = tmp_path / "legacy_nofilter_2.xls"
    legacy_candidate.write_bytes(b"legacy temporary")
    messages = []

    removed_count = file_runtime.cleanup_nofilter_files(
        log_func=messages.append,
        work_dir=str(tmp_path),
    )

    assert removed_count == 2
    assert not candidate.exists()
    assert not legacy_candidate.exists()
    assert any("已清理 nofilter" in message for message in messages)


def test_cleanup_nofilter_files_logs_remove_failure(
    monkeypatch, tmp_path: Path
) -> None:
    candidate = tmp_path / "source_nofilter_1.xlsx"
    candidate.write_bytes(b"temporary")
    messages = []

    def fail_remove(_path):
        raise OSError("remove failed")

    monkeypatch.setattr(file_runtime.os, "remove", fail_remove)

    removed_count = file_runtime.cleanup_nofilter_files(
        log_func=messages.append,
        work_dir=str(tmp_path),
    )

    assert removed_count == 0
    assert any("删除临时缓存文件失败" in message for message in messages)


def test_cleanup_nofilter_files_handles_missing_dir(tmp_path: Path) -> None:
    missing_dir = tmp_path / "missing"
    removed_count = file_runtime.cleanup_nofilter_files(
        work_dir=str(missing_dir),
    )
    assert removed_count == 0


def test_cleanup_nofilter_files_logs_outer_failure(monkeypatch) -> None:
    messages = []

    def fail_temp_dir():
        raise RuntimeError("temp failed")

    monkeypatch.setattr(file_runtime, "get_app_temp_dir", fail_temp_dir)

    assert file_runtime.cleanup_nofilter_files(messages.append) == 0
    assert any("清理临时缓存文件时出错" in message for message in messages)


def _write_workbook(path: Path) -> Path:
    workbook = Workbook()
    workbook.active.title = "Sheet"
    workbook.active.append(["id", "value"])
    workbook.save(path)
    workbook.close()
    return path


def _raise_value_error():
    raise ValueError("workbook unavailable")
