# -*- coding: utf-8 -*-
"""增强比对控制与输出排序（include_sheets / ignore_cols / sheet_order）集成测试。"""

import os
import threading

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.backend.domain.data_comparison import process_edc_multithreaded
from src.backend.infrastructure.config_manager import ConfigManager


def _make_wb(path, sheets):
    wb = Workbook()
    wb.remove(wb.active)
    for name, df in sheets.items():
        ws = wb.create_sheet(title=name)
        ws.append(list(df.columns))
        for row in df.itertuples(index=False):
            ws.append(list(row))
    wb.save(path)


def _run(old_path, new_path, out_path, params):
    params = dict({"default_keys": ["SUBJID"]}, **params)
    cm = ConfigManager()
    cm.update_from_parameters(params, {})
    result = process_edc_multithreaded(
        old_path,
        new_path,
        out_path,
        lambda m: None,
        config=cm,
        stop_flag=threading.Event(),
    )
    return load_workbook(result)


@pytest.fixture
def workbooks(tmp_path):
    """构造新旧文件：AE(AEMODIFY+UPDATETIME 变化) / DM(仅 UPDATETIME 变化) / CM(无变化)。"""
    old_path = os.path.join(str(tmp_path), "old.xlsx")
    new_path = os.path.join(str(tmp_path), "new.xlsx")

    df_ae_old = pd.DataFrame(
        {"SUBJID": ["001"], "AEMODIFY": ["a"], "UPDATETIME": ["2026-01-01"]}
    )
    df_ae_new = pd.DataFrame(
        {"SUBJID": ["001"], "AEMODIFY": ["b"], "UPDATETIME": ["2026-08-18"]}
    )
    df_dm_old = pd.DataFrame(
        {"SUBJID": ["001"], "AGE": [30], "UPDATETIME": ["2026-01-01"]}
    )
    df_dm_new = pd.DataFrame(
        {"SUBJID": ["001"], "AGE": [30], "UPDATETIME": ["2026-08-18"]}
    )
    df_cm_old = pd.DataFrame({"SUBJID": ["001"], "CMTERM": ["aspirin"]})
    df_cm_new = pd.DataFrame({"SUBJID": ["001"], "CMTERM": ["aspirin"]})

    # 新文件 sheet 顺序 AE, DM, CM（与字母序 AE,CM,DM 不同，用于验证排序）
    _make_wb(old_path, {"AE": df_ae_old, "DM": df_dm_old, "CM": df_cm_old})
    _make_wb(new_path, {"AE": df_ae_new, "DM": df_dm_new, "CM": df_cm_new})
    return old_path, new_path


def _mark(wb, sheet):
    ws = wb[sheet]
    headers = [c.value for c in ws[1]]
    mark_col = headers.index("更新情况（标记）") + 1
    return ws.cell(row=2, column=mark_col).value


def _headers(wb, sheet):
    return [c.value for c in wb[sheet][1]]


@pytest.mark.integration
def test_include_sheets_and_order(workbooks, tmp_path):
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    wb = _run(
        old_path,
        new_path,
        out_path,
        {
            "include_sheets": ["DM", "AE"],
            "sheet_order": ["AE", "DM"],
        },
    )
    assert wb.sheetnames == ["比对结果汇总", "AE", "DM"]


@pytest.mark.integration
def test_default_order_follows_source_file(workbooks, tmp_path):
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    wb = _run(old_path, new_path, out_path, {})
    assert wb.sheetnames[0] == "比对结果汇总"
    # 新文件 sheet 顺序: AE, DM, CM
    assert wb.sheetnames[1:] == ["AE", "DM", "CM"]


@pytest.mark.integration
def test_global_ignore_cols(workbooks, tmp_path):
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    wb = _run(
        old_path,
        new_path,
        out_path,
        {
            "ignore_cols": ["UPDATETIME"],
            "include_sheets": ["AE", "DM"],
        },
    )
    # AE 的 AEMODIFY 变化未被忽略 → 更新
    assert _mark(wb, "AE") == "更新"
    # DM 仅 UPDATETIME 变化且被忽略 → 未改变
    assert _mark(wb, "DM") == "未改变"
    # 被忽略列照常输出
    dm_headers = [c.value for c in wb["DM"][1]]
    assert "UPDATETIME" in dm_headers


@pytest.mark.integration
def test_sheet_ignore_cols_replaces_global(workbooks, tmp_path):
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    wb = _run(
        old_path,
        new_path,
        out_path,
        {
            "ignore_cols": ["UPDATETIME"],
            "sheet_ignore_cols": {"AE": ["AEMODIFY"]},
            "include_sheets": ["AE", "DM"],
        },
    )
    # AE 整体替换为忽略 AEMODIFY（UPDATETIME 不再被忽略）→ 剩余 UPDATETIME 变化 → 更新
    # （若错误实现为叠加，则 AE 应为未改变）
    assert _mark(wb, "AE") == "更新"
    # DM 无 per-sheet 覆盖 → 全局忽略 UPDATETIME 仍生效 → 未改变
    assert _mark(wb, "DM") == "未改变"


@pytest.mark.integration
def test_global_common_cols_drops_columns(workbooks, tmp_path):
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    wb = _run(
        old_path,
        new_path,
        out_path,
        {
            "common_cols": ["UPDATETIME"],
            "include_sheets": ["AE", "DM"],
        },
    )
    # 排除字段在读取期物理删列：AE/DM 表头均无该列
    assert "UPDATETIME" not in _headers(wb, "AE")
    assert "UPDATETIME" not in _headers(wb, "DM")
    # 未排除的列照常输出
    assert "AEMODIFY" in _headers(wb, "AE")
    # 排除 UPDATETIME 后 AE 仍有 AEMODIFY 变化 → 更新；DM 无剩余变化 → 未改变
    assert _mark(wb, "AE") == "更新"
    assert _mark(wb, "DM") == "未改变"


@pytest.mark.integration
def test_sheet_common_cols_replaces_global(workbooks, tmp_path):
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    wb = _run(
        old_path,
        new_path,
        out_path,
        {
            "common_cols": ["UPDATETIME"],
            "sheet_common_cols": {"AE": ["AEMODIFY"]},
            "include_sheets": ["AE", "DM"],
        },
    )
    # AE 整体替换为排除 AEMODIFY（UPDATETIME 不再被排除）
    # （若错误实现为叠加，UPDATETIME 会被一起删除，此断言失败）
    assert "AEMODIFY" not in _headers(wb, "AE")
    assert "UPDATETIME" in _headers(wb, "AE")
    # 剩余 UPDATETIME 有变化 → 更新
    assert _mark(wb, "AE") == "更新"
    # DM 未命中 per-sheet → 回退全局排除 UPDATETIME
    assert "UPDATETIME" not in _headers(wb, "DM")


@pytest.mark.integration
def test_sheet_common_cols_empty_list_disables_global(workbooks, tmp_path):
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    wb = _run(
        old_path,
        new_path,
        out_path,
        {
            "common_cols": ["UPDATETIME"],
            "sheet_common_cols": {"AE": []},
            "include_sheets": ["AE", "DM"],
        },
    )
    # 空列表 = 该表单什么都不删（若误用真值判断，会静默回退全局排除）
    assert "UPDATETIME" in _headers(wb, "AE")
    assert "AEMODIFY" in _headers(wb, "AE")
    # DM 未命中 per-sheet → 仍按全局排除
    assert "UPDATETIME" not in _headers(wb, "DM")


@pytest.mark.integration
def test_sheet_common_cols_anchor_warning_when_global_has_key(workbooks, tmp_path):
    """显式空列表禁用全局排除时，全局列表里的锚点列仍会先被删掉 → 应打锚点失效告警。"""
    old_path, new_path = workbooks
    out_path = os.path.join(str(tmp_path), "out.xlsx")
    warnings = []

    def collect_log(message):
        warnings.append(message)

    params = dict(
        {
            "default_keys": ["SUBJID"],
            "common_cols": ["SUBJID"],
            "sheet_common_cols": {"AE": []},
            "include_sheets": ["AE", "DM"],
        },
    )
    cm = ConfigManager()
    cm.update_from_parameters(params, {})
    process_edc_multithreaded(
        old_path,
        new_path,
        out_path,
        collect_log,
        config=cm,
        stop_flag=threading.Event(),
    )
    assert any("锚点将失效" in message for message in warnings)


def test_config_manager_defaults_sheet_common_cols():
    cm = ConfigManager()
    assert cm.sheet_common_cols == {}
    # 老配置缺 sheet_common_cols 键 → 回退空字典，仅全局生效
    cm.update_from_parameters({"common_cols": ["A"]}, {})
    assert cm.sheet_common_cols == {}
    assert cm.common_cols_to_drop == ["A"]
