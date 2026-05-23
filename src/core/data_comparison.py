import pandas as pd
from numbers import Number
import gc
import threading
import os
import concurrent.futures
from openpyxl import Workbook,load_workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Border, Side, Font, Alignment

# 导入psutil用于内存监控
import psutil

# 导入重构后的模块
from ..utils.progress_manager import ThreadSafeProgressManager
from ..utils.sheet_process_result import SheetProcessResult
from ..utils.gui_update_manager import GUIUpdateType, GUIUpdateMessage, GUIUpdateManager
from ..utils.file_utils import check_and_remove_file_protection, validate_excel_file, get_resource_path, set_window_icon, remove_auto_filters_from_xlsx, reorder_columns_with_update_mark_first, get_sheet_names, get_app_temp_dir, cleanup_nofilter_files
from ..utils.logger import log
from ..utils.stop_flag_utils import check_stop_frequently, check_stop, set_global_stop_flag, update_progress
from .excel_utils import replace_worksheet_headers, apply_highlight_to_worksheet
from ..utils.config_manager import ConfigManager
from .excel_header_utils import read_single_sheet_from_excel

def _maybe_gc_collect(threshold_percent=70, log_func=None):
    """
    根据内存使用情况有条件地执行垃圾回收。
    需要安装 psutil 库（pip install psutil）。
    如果 psutil 未安装或出错，将退化为无条件执行 gc.collect()。
    """
    try:
        process = psutil.Process(os.getpid())
        current_process_usage_percent = process.memory_percent()
        system_overall_usage_percent = psutil.virtual_memory().percent

        print(f"当前进程内存使用率: {current_process_usage_percent:.2f}% | 系统总内存使用率: {system_overall_usage_percent:.2f}% (GC阈值: {threshold_percent}%)")

        if current_process_usage_percent >= threshold_percent:
            if log_func:
                log_func(f"内存使用达到 {current_process_usage_percent:.2f}%，执行垃圾回收。")
            gc.collect()
            if log_func:
                log_func(f"垃圾回收完成，当前进程内存使用率: {process.memory_percent():.2f}%。")
    except ImportError:
        if log_func:
            log_func("警告: psutil 库未安装。无法执行基于内存的垃圾回收，将无条件执行 gc.collect()。请运行 'pip install psutil' 安装。")
        gc.collect()
    except Exception as e:
        if log_func:
            log_func(f"执行条件垃圾回收时发生错误: {str(e)}。将执行无条件垃圾回收。")
        gc.collect()

def process_single_sheet_complete(sheet_name, old_path, new_path, config, progress_manager):
    """
    处理单个Excel Sheet的完整流程：包括读取数据、比对差异，并返回处理结果。
    此函数设计为在独立线程中运行，每个线程负责一个Sheet的处理，并独立地打开和关闭文件，
    以确保资源隔离和多线程效率。
    
    Args:
        sheet_name (str): 当前要处理的Sheet名称。
        old_path (str): 旧版本Excel文件路径。
        new_path (str): 新版本Excel文件路径。
        config (ConfigManager): 包含各种比对和处理配置的对象。
        progress_manager (ThreadSafeProgressManager): 线程安全的进度管理器，用于更新GUI进度和日志。
        
    Returns:
        SheetProcessResult: 包含处理结果DataFrame的SheetProcessResult对象。
    """
    result = SheetProcessResult(sheet_name) # 初始化处理结果对象
    log_func = progress_manager.safe_log # 获取线程安全的日志函数

    # 若当前sheet在排除列表中，直接跳过处理
    try:
        exclude_sheets = getattr(config, 'exclude_sheets', []) or []
        if sheet_name in exclude_sheets:
            log(f"⏭️ 跳过Sheet [{sheet_name}]：已在排除列表中", log_func)
            result.success = True
            progress_manager.update_sheet_progress(sheet_name, "跳过")
            return result
    except Exception:
        pass

    try:
        # print(f"开始处理：[{sheet_name}]")
        progress_manager.update_sheet_progress(sheet_name, "正在处理") # 更新GUI进度

        check_stop_frequently(progress_manager.log_func) # 频繁检查是否需要停止操作
        
        anchor_row_num = config.anchor_row_num
        header_row_num = config.header_row_num

        # 1. 读取旧版本和新版本文件中的单个Sheet数据
        # log_func(f"读取旧版本文件：[{sheet_name}]")
        # read_single_sheet_from_excel 返回 DataFrame 和 None（因为不再需要原始工作表对象进行格式复制）
        old_df_raw = read_single_sheet_from_excel(old_path, sheet_name, anchor_row_num, header_row_num, log_func, config.common_cols_to_drop) 
        new_df_raw = read_single_sheet_from_excel(new_path, sheet_name, anchor_row_num, header_row_num, log_func, config.common_cols_to_drop)

        old_df = old_df_raw if old_df_raw is not None else None
        new_df = new_df_raw if new_df_raw is not None else None

        # 显式释放原始DataFrame，避免内存峰值
        del old_df_raw
        del new_df_raw
        # 将 gc.collect() 替换为条件垃圾回收
        _maybe_gc_collect(threshold_percent=70, log_func=log_func) 

        # 确定原始源文件，用于后续复制格式（优先新文件，其次旧文件）
        if new_df is not None:
            result.original_source_file = new_path
            result.original_source_sheet_name = sheet_name
        elif old_df is not None:
            result.original_source_file = old_path
            result.original_source_sheet_name = sheet_name

        # === 修正：先判断None，再判断empty，避免NoneType异常 ===
        if old_df is None and new_df is None:
            # 如果Sheet在新旧文件中都不存在，则跳过处理
            log_func(f"ℹ️ Sheet [{sheet_name}] 在新旧文件中均不存在，跳过比对")
            result.success = True 
            return result
        elif old_df is None and new_df is not None:
            # 如果旧文件中不存在但新文件中存在，标记为新增Sheet
            log_func(f"ℹ️ 检测到新增Sheet: [{sheet_name}]，跳过比对")
            progress_manager.update_sheet_progress(sheet_name, "新增表单") # 更新进度状态
            
            # 处理新增Sheet，返回带有'新增'标记的DataFrame
            output_df = process_new_sheet(sheet_name, new_df, config, progress_manager)
            result.change_type = "new" # 设置变更类型为新增
            result.differences = None # 新增表单没有单元格级别的差异
            result.success = True
            result.df = output_df # 直接存储最终的DataFrame
            # 拷贝SAS元数据从DataFrame到result对象
            if hasattr(output_df, 'attrs'):
                result.sas_file_names = output_df.attrs.get('sas_file_names', [])
                result.sas_file_labels = output_df.attrs.get('sas_file_labels', [])
                result.sas_name_to_label = output_df.attrs.get('sas_name_to_label', {})
            # 统计计数
            try:
                if '更新情况（标记）' in output_df.columns:
                    result.added_rows_count = int((output_df['更新情况（标记）'] == '新增').sum())
            except Exception:
                result.added_rows_count = 0
            
            print(f"✅ Sheet [{sheet_name}] 处理完成。")
            return result
        elif old_df is not None and new_df is None:
            # 如果新文件中不存在但旧文件中存在，标记为缺失Sheet（删除）
            log_func(f"ℹ️ 检测到缺失Sheet: [{sheet_name}]，跳过比对") # 修正日志信息，更简洁
            progress_manager.update_sheet_progress(sheet_name, "缺失表单") # 更新进度状态
            
            # 如果不合并删除数据，则直接跳过写入该缺失Sheet
            if hasattr(config, 'merge_deleted_data') and not config.merge_deleted_data:
                result.change_type = None
                result.differences = None
                result.success = True
                result.df = None
                return result

            # 处理缺失Sheet，返回带有'删除'标记的DataFrame
            output_df = process_missing_sheet(sheet_name, old_df, config, progress_manager)
            result.change_type = "missing" # 设置变更类型为缺失
            result.differences = None # 缺失表单没有单元格级别的差异
            result.success = True
            result.df = output_df # 直接存储最终的DataFrame
            # 拷贝SAS元数据从DataFrame到result对象
            if hasattr(output_df, 'attrs'):
                result.sas_file_names = output_df.attrs.get('sas_file_names', [])
                result.sas_file_labels = output_df.attrs.get('sas_file_labels', [])
                result.sas_name_to_label = output_df.attrs.get('sas_name_to_label', {})
            # 统计计数
            try:
                if '更新情况（标记）' in output_df.columns:
                    result.deleted_rows_count = int((output_df['更新情况（标记）'] == '删除').sum())
            except Exception:
                result.deleted_rows_count = 0

            print(f"✅ Sheet [{sheet_name}] 处理完成。")
            return result
        # 修正：将空的DataFrame情况提升到这里单独处理，确保生成结果
        elif old_df is not None and new_df is not None and old_df.empty and new_df.empty:
            log_func(f"ℹ️ Sheet [{sheet_name}] 在新旧版本中均为空。")
            # 创建一个仅包含'更新情况（标记）'的空DataFrame
            output_df = pd.DataFrame(columns=['更新情况（标记）'])
            
            # 尝试合并新旧DF的列名和列标签，以提供更完整的空表头
            all_original_cols = list(set(old_df.columns).union(set(new_df.columns))) # 获取所有唯一列
            if '更新情况（标记）' in all_original_cols: # 确保它不会被重复添加
                all_original_cols.remove('更新情况（标记）')
            all_original_cols.insert(0, '更新情况（标记）') # 始终将标记列放在第一位
            
            merged_name_to_label = {}
            if hasattr(old_df, 'attrs'):
                merged_name_to_label.update(old_df.attrs.get('sas_name_to_label', {}))
            if hasattr(new_df, 'attrs'):
                merged_name_to_label.update(new_df.attrs.get('sas_name_to_label', {}))
            merged_name_to_label['更新情况（标记）'] = '更新情况（标记）' # 确保标记列的标签

            output_df.attrs['sas_file_names'] = all_original_cols
            output_df.attrs['sas_file_labels'] = [merged_name_to_label.get(col, col) for col in all_original_cols]
            output_df.attrs['sas_name_to_label'] = merged_name_to_label
            output_df.attrs['change_type'] = None # 无实际变化
            
            result.success = True
            result.change_type = None
            result.differences = {} # 空表单没有单元格差异
            result.df = output_df
            print(f"✅ Sheet [{sheet_name}] 处理完成。")
            return result
        # 下面才是正常比对流程：新旧文件都存在且不为空
        elif old_df is not None and new_df is not None:
            # 检查 old_df 是否为空DataFrame，且 new_df 不为空
            if hasattr(old_df, 'empty') and old_df.empty and not (hasattr(new_df, 'empty') and new_df.empty):
                log_func(f"ℹ️ Sheet [{sheet_name}] (旧版) 没有数据，新版有数据，视为新增内容。")
                # 使用process_new_sheet生成带有"新增"标记和new_df表头的DataFrame
                output_df = process_new_sheet(sheet_name, new_df, config, progress_manager)
                
                result.change_type = "new"
                result.differences = None # 无单元格差异，因为是新增内容
                result.success = True
                result.df = output_df
                
                # 拷贝SAS元数据从output_df到result对象
                if hasattr(output_df, 'attrs'):
                    result.sas_file_names = output_df.attrs.get('sas_file_names', [])
                    result.sas_file_labels = output_df.attrs.get('sas_file_labels', [])
                    result.sas_name_to_label = output_df.attrs.get('sas_name_to_label', {})
                # 统计计数
                try:
                    if '更新情况（标记）' in output_df.columns:
                        result.added_rows_count = int((output_df['更新情况（标记）'] == '新增').sum())
                except Exception:
                    result.added_rows_count = 0
                
                print(f"✅ Sheet [{sheet_name}] 处理完成。")
                return result
            # 判断新DF为空但旧DF不为空的情况（视为内容缺失）
            elif new_df.empty and not old_df.empty:
                log_func(f"ℹ️ Sheet [{sheet_name}] (新版) 没有数据，旧版有数据，视为缺失内容。")
                output_df = process_missing_sheet(sheet_name, old_df, config, progress_manager)
                
                result.change_type = "missing"
                result.differences = None # 无单元格差异，因为是缺失内容
                result.success = True
                result.df = output_df
                
                # 拷贝SAS元数据从output_df到result对象
                if hasattr(output_df, 'attrs'):
                    result.sas_file_names = output_df.attrs.get('sas_file_names', [])
                    result.sas_file_labels = output_df.attrs.get('sas_file_labels', [])
                    result.sas_name_to_label = output_df.attrs.get('sas_name_to_label', {})

                print(f"✅ Sheet [{sheet_name}] 处理完成。")
                # 统计计数
                try:
                    if '更新情况（标记）' in output_df.columns:
                        result.deleted_rows_count = int((output_df['更新情况（标记）'] == '删除').sum())
                except Exception:
                    result.deleted_rows_count = 0
                return result
            
            # 如果都不是上述特殊情况（即新旧DF都有数据且都不为空），则进行正常比对
            else:
                # log_func(f"正在比对：[{sheet_name}]")
                progress_manager.update_sheet_progress(sheet_name, "正在比对") # 更新进度状态
                
                # 确定当前Sheet使用的锚点列
                if sheet_name in config.sheet_key_map:
                    key_cols = config.sheet_key_map[sheet_name]
                    log_func(f"ℹ️ Sheet [{sheet_name}] 使用指定锚点列: {key_cols}")
                else:
                    key_cols = config.default_keys
                    log_func(f"ℹ️ Sheet [{sheet_name}] 使用默认锚点列: {key_cols}")
                
                # 执行完整的比对逻辑，返回比对后的DataFrame和差异信息
                compared_df, diffs, added_cols, deleted_cols, sas_file_names, sas_file_labels, sas_name_to_label, change_type, updated_count, deleted_count, added_count = \
                    perform_full_comparison(sheet_name, old_df, new_df, key_cols, config, progress_manager)
                
                result.change_type = change_type # 设置变更类型
                result.differences = diffs # 存储差异信息
                result.add_sas_names = added_cols # 存储新增列
                result.del_sas_names = deleted_cols # 存储删除列
                result.sas_file_names = sas_file_names # 存储最终列名
                result.sas_file_labels = sas_file_labels # 存储最终列标签
                result.sas_name_to_label = sas_name_to_label # 存储列名到标签的映射
                result.df = compared_df # 存储最终比对结果DataFrame
                result.success = True # 标记处理成功
                # 计数赋值
                result.updated_rows_count = int(updated_count or 0)
                result.deleted_rows_count = int(deleted_count or 0)
                result.added_rows_count = int(added_count or 0)
                
                print(f"✅ Sheet [{sheet_name}] 处理完成。")
                return result
        else:
            # 理论上不会走到这里，除非文件存在性判断逻辑有误
            log_func(f"❓ Sheet [{sheet_name}] 状态异常，未能处理。")
            result.success = False
            result.error_message = "未知的Sheet状态" # 未知错误状态
            return result
    except InterruptedError:
        raise # 重新抛出停止信号，以便主线程捕获并终止任务
    except Exception as e:
        error_msg = f"处理Sheet [{sheet_name}] 时出错: {str(e)}"
        log_func(f"❌ {error_msg}") # 记录错误日志
        result.success = False
        result.error_message = error_msg # 存储错误信息
        return result
    finally:
        # 无论成功失败，都更新该Sheet的最终进度
        progress_manager.update_sheet_progress(sheet_name, "完成" if result.success else "失败", is_final_update=True)
        # 将 gc.collect() 替换为条件垃圾回收
        _maybe_gc_collect(threshold_percent=70, log_func=log_func) 

def process_missing_sheet(sheet_name, old_df, config, progress_manager):
    """处理缺失的表单（即旧文件中存在，新文件中不存在的表单）。
    会给缺失的行添加"删除"标记，并根据配置删除指定列。
    Args:
        sheet_name (str): 缺失的Sheet名称。
        old_df (pd.DataFrame): 旧版本Sheet的DataFrame数据。
        config (ConfigManager): 配置对象。
        progress_manager (ThreadSafeProgressManager): 线程安全的进度管理器。
    Returns:
        pd.DataFrame: 处理后的DataFrame，包含"更新情况（标记）"列和SAS元数据。
    """
    try:
        missing_df = old_df
        missing_df['更新情况（标记）'] = '删除' # 添加标记列并设为"删除"
        missing_df = reorder_columns_with_update_mark_first(missing_df) # 将标记列移到第一列
        
        # 从DataFrame的attrs中获取SAS列名到标签的映射
        sas_name_to_label = getattr(missing_df, 'attrs', {}).get('sas_name_to_label', {})
               
        if not hasattr(missing_df, 'attrs'):
            missing_df.attrs = {}
        missing_df.attrs['sas_name_to_label'] = sas_name_to_label
        
        # 生成表头和数据元数据，这些信息将存储在DataFrame的attrs中
        all_missing_columns = list(missing_df.columns) # 获取所有列名
        complete_missing_labels = [] # 存储完整的列标签
        for col in all_missing_columns:
            if col == '更新情况（标记）':
                complete_missing_labels.append('更新情况（标记）')
            else:
                # 从sas_name_to_label获取对应标签，如果没有则使用列名本身
                complete_missing_labels.append(sas_name_to_label.get(col, col))
        
        # 将这些元数据存储在DataFrame的attrs中，供后续使用
        missing_df.attrs['sas_file_names'] = all_missing_columns
        missing_df.attrs['sas_file_labels'] = complete_missing_labels
        missing_df.attrs['sas_name_to_label'] = sas_name_to_label # 再次确保映射被存储
        missing_df.attrs['change_type'] = 'missing' # 设置变更类型
        
        return missing_df
        
    except Exception as e:
        progress_manager.safe_log(f"处理缺失Sheet [{sheet_name}] 时出错: {str(e)}") # 记录错误日志
        return pd.DataFrame() # 出错时返回空DataFrame

def process_new_sheet(sheet_name, new_df, config, progress_manager):
    """处理新增的表单（即旧文件中不存在，新文件中存在的表单）。
    会给新增的行添加"新增"标记，并根据配置删除指定列。
    Args:
        sheet_name (str): 新增的Sheet名称。
        new_df (pd.DataFrame): 新版本Sheet的DataFrame数据。
        config (ConfigManager): 配置对象。
        progress_manager (ThreadSafeProgressManager): 线程安全的进度管理器。
    Returns:
        pd.DataFrame: 处理后的DataFrame，包含"更新情况（标记）"列和SAS元数据。
    """
    try:
        # No copy to reduce memory footprint; this path does not reuse the input DataFrame.
        new_df['更新情况（标记）'] = '' # 初始设置为空字符串，待后续判断是否实际有数据
        new_df = reorder_columns_with_update_mark_first(new_df) # 将标记列移到第一列
        
        # 从DataFrame的attrs中获取SAS列名到标签的映射
        new_name_to_label = getattr(new_df, 'attrs', {}).get('sas_name_to_label', {})
                
        # 生成表头元数据，这些信息将存储在DataFrame的attrs中
        all_columns = list(new_df.columns) # 获取所有列名
        complete_sas_labels = [] # 存储完整的列标签
        for col in all_columns:
            if col == '更新情况（标记）':
                complete_sas_labels.append('更新情况（标记）')
            else:
                # 从new_name_to_label获取对应标签，如果没有则使用列名本身
                complete_sas_labels.append(new_name_to_label.get(col, col))
        
        # 处理数据：如果DataFrame不为空，则将所有行的"更新情况（标记）"设为"新增"
        if len(new_df) == 0:
            progress_manager.safe_log(f"新增表单[{sheet_name}]无任何数据")
        else:
            new_df['更新情况（标记）'] = '新增'
        
        # 将这些元数据存储在DataFrame的attrs中，供后续使用
        new_df.attrs['sas_file_names'] = all_columns
        new_df.attrs['sas_file_labels'] = complete_sas_labels
        new_df.attrs['sas_name_to_label'] = new_name_to_label # 再次确保映射被存储
        new_df.attrs['change_type'] = 'new' # 设置变更类型

        return new_df
                
    except Exception as e:
        progress_manager.safe_log(f"处理新增Sheet [{sheet_name}] 时出错: {str(e)}") # 记录错误日志
        return pd.DataFrame() # 出错时返回空DataFrame

def perform_full_comparison(sheet_name, old_df, new_df, key_cols, config, progress_manager):
    """
    执行完整的数据比对流程。
    主要步骤包括：创建锚点、识别列的增删、合并数据、执行单元格级别比对，
    并根据比对结果更新"更新情况（标记）"列，最后清理辅助列并返回结果。
    
    Args:
        sheet_name (str): 当前处理的Sheet名称。
        old_df (pd.DataFrame): 旧版本Sheet的DataFrame数据。
        new_df (pd.DataFrame): 新版本Sheet的DataFrame数据。
        key_cols (list): 用于创建锚点（唯一标识行）的列名列表。
        config (ConfigManager): 配置管理器对象，包含颜色等设置。
        progress_manager (ThreadSafeProgressManager): 线程安全的进度管理器，用于日志记录。
    
    Returns:
        tuple: 包含以下元素的元组
            - compared_df (pd.DataFrame): 最终比对结果DataFrame。
            - diffs (dict): 单元格级别的差异信息（格式为 {row_idx: {col_name: flag}}）。
            - added_cols (list): 新增的列名列表。
            - deleted_cols (list): 删除的列名列表。
            - sas_file_names (list): 最终结果DataFrame的列名列表。
            - sas_file_labels (list): 最终结果DataFrame的列标签列表。
            - sas_name_to_label (dict): 列名到标签的映射字典。
            - change_type (str): Sheet的总体变更类型（'data_changed'或None）。
            - updated_count (int): 行级"更新"的数量。
            - deleted_count (int): 行级"删除"的数量。
            - added_count (int): 行级"新增"的数量。
    """
    try:       
        # print(f"   [比对逻辑开始] Sheet: {sheet_name}")
        # print(f"   使用的锚点列: {key_cols}")

        # 创建锚点列(_ANCHOR)用于行匹配
        new_df = create_anchor_by_sas_names(new_df, key_cols, progress_manager.safe_log, sheet_name)
        old_df = create_anchor_by_sas_names(old_df, key_cols, progress_manager.safe_log, sheet_name)
        
        # 打印锚点信息
        # print(f"   新DF锚点（前5行）:\n{new_df['_ANCHOR'].head().to_string()}")
        # print(f"   旧DF锚点（前5行）:\n{old_df['_ANCHOR'].head().to_string()}")

        # 在合并 DataFrame 之前，基于原始列名识别新增和删除的列
        # 获取旧DF的原始SAS列名，如果attrs中没有，则使用DF的列名
        old_original_cols_names = set(getattr(old_df, 'attrs', {}).get('sas_file_name', list(old_df.columns)))
        # 获取新DF的原始SAS列名
        new_original_cols_names = set(getattr(new_df, 'attrs', {}).get('sas_file_name', list(new_df.columns)))
        
        # 计算新增的列：在新DF有但在旧DF没有的列
        added_cols = [col for col in new_original_cols_names if col not in old_original_cols_names]
        # 计算删除的列：在旧DF有但在新DF没有的列
        deleted_cols = [col for col in old_original_cols_names if col not in new_original_cols_names]

        # 合并列标签映射：优先使用新DF的标签，旧DF独有的标签也加入
        new_name_to_label = getattr(new_df, 'attrs', {}).get('sas_name_to_label', {})
        old_name_to_label = getattr(old_df, 'attrs', {}).get('sas_name_to_label', {})
        
        merged_name_to_label = new_name_to_label.copy() # 以新DF的映射为基础
        for old_name, old_label in old_name_to_label.items():
            if old_name not in merged_name_to_label:
                merged_name_to_label[old_name] = old_label # 将旧DF独有的映射添加进来
            # 为旧版本列名（带_OLD_后缀）也添加映射，方便后续获取其标签
            old_name_with_suffix = f"{old_name}_OLD_"
            merged_name_to_label[old_name_with_suffix] = old_label
        
        # 合并数据：使用锚点列进行外连接，保留所有行
        merged_df = pd.merge(
            new_df, 
            old_df, 
            on='_ANCHOR', # 合并键为锚点列
            how='outer', # 外连接，保留所有新旧行的记录
            suffixes=('', '_OLD_') # 旧版本列添加_OLD_后缀
        )

        merged_df['更新情况（标记）'] = '未改变' # 初始化"更新情况（标记）"列为"未改变"

        # 打印合并后DataFrame的锚点信息
        # print(f"   合并后DF锚点（前5行）:\n{merged_df['_ANCHOR'].head().to_string()}")

        # 设置合并后DataFrame的attrs，特别是sas_name_to_label
        if not hasattr(merged_df, 'attrs'):
            merged_df.attrs = {}
        merged_df.attrs['sas_name_to_label'] = merged_name_to_label
        
        # 执行比对，找出单元格级别的差异
        dift = compare_columns_by_sas_names( # 确保这里调用的是正确的函数
            merged_df=merged_df, # 传入合并后的DataFrame
            key_sas_names=key_cols, # 锚点列，比对时会忽略
            sheet_name=sheet_name,
            log_func=progress_manager.safe_log
        )
        
        has_changes = False # 标记整个Sheet是否有数据变化

        # 处理变动信息：根据dift结果更新"更新情况（标记）"列和diff_dict
        if not dift.empty:
            has_changes = True # 存在差异，设置标志为True
            
            # 处理行变动情况：聚合同一行的所有差异标记
            row_changes = {} # 存储每行变化的标记，如 {'row_idx': ['新增', '更新', '删除']} 
            diff_dict = {} # 存储单元格级别的差异详情，格式为 {row_idx: {col_name: flag}}
            for idx, row in dift.iterrows(): # 遍历dift DataFrame的每一行
                try:
                    row_idx = row['row'] # 获取行索引
                    col_name = row['col'] # 获取列名
                    flag = row['flag'] # 获取差异标记（新增/删除/更新）
  
                    # 确保row_idx是有效的，跳过NaN或None
                    if pd.isna(row_idx) or row_idx is None:
                        continue

                    if row_idx not in diff_dict:
                        diff_dict[row_idx] = {} # 如果是新行，初始化其差异字典
                    diff_dict[row_idx][col_name] = flag # 记录单元格差异

                    # 初始化 row_changes
                    if row_idx not in row_changes:
                        row_changes[row_idx] = []
                    # 确保列表不为None（防御性编程）
                    if row_changes[row_idx] is None:
                        row_changes[row_idx] = []
                        
                    row_changes[row_idx].append(flag)
                except Exception as e:
                    progress_manager.safe_log(f"处理行变动信息时出错: {str(e)}, idx={idx}")
                    continue # 跳过当前行，继续处理下一行
            
            # print(f"   Sheet [{sheet_name}] diff_dict（前5条）:\n{dict(list(diff_dict.items())[:5])}")

            # 更新"更新情况（标记）"列
            for row_idx, changes in row_changes.items():
                is_new_record = True # 假设是新增记录
                is_deleted_record = True # 假设是删除记录
                
                # 检查新版本行中是否有数据（通过遍历新DF的原始列来判断）
                for col in new_original_cols_names:
                    if col in merged_df.columns and pd.notna(merged_df.loc[row_idx, col]) and str(merged_df.loc[row_idx, col]).strip() != '':
                        is_deleted_record = False # 发现新数据，则不是纯删除记录
                        break
                
                # 检查旧版本行中是否有数据（通过遍历旧DF的原始列来判断）
                for col in old_original_cols_names:
                    old_col_name = f"{col}_OLD_"
                    if old_col_name in merged_df.columns and pd.notna(merged_df.loc[row_idx, old_col_name]) and str(merged_df.loc[row_idx, old_col_name]).strip() != '':
                        is_new_record = False # 发现旧数据，则不是纯新增记录
                        break
                
                # print(f"   行 {row_idx}: is_new_record={is_new_record}, is_deleted_record={is_deleted_record}, changes={changes}")

                if is_new_record and not is_deleted_record: 
                    merged_df.loc[row_idx, "更新情况（标记）"] = "新增" # 标记为新增记录
                elif is_deleted_record and not is_new_record: 
                    merged_df.loc[row_idx, "更新情况（标记）"] = "删除" # 标记为删除记录
                    # 仅当需要合并删除数据时，才将旧值恢复到主列
                    if getattr(config, 'merge_deleted_data', True):
                        for col_name in old_original_cols_names:
                            old_field_name = f"{col_name}_OLD_"
                            if old_field_name in merged_df.columns and col_name in merged_df.columns:
                                merged_df.loc[row_idx, col_name] = merged_df.loc[row_idx, old_field_name]
                else:
                    # 如果不是纯新增或纯删除，则检查是否有单元格级别的更新
                    has_actual_cell_changes = False
                    for change_flag in changes: 
                        if change_flag in ['新增', '删除', '更新']: # 只要有这三种标记，就认为有实际变化
                            has_actual_cell_changes = True
                            break
                    
                    if has_actual_cell_changes:
                        merged_df.loc[row_idx, "更新情况（标记）"] = "更新"
                    else:
                        # 没有实际单元格变化，保持"未改变"状态
                        merged_df.loc[row_idx, "更新情况（标记）"] = "未改变"
        
        else:
            # 如果dift为空，表示没有单元格级别的差异，则整个Sheet标记为"未改变"
            # progress_manager.safe_log(f"ℹ️ 表单 {sheet_name} 无差异。")
            merged_df['更新情况（标记）'] = '未改变' # 确保标记为未改变
            diff_dict = {} # 无差异，diff_dict为空
        
        # 在删除临时列和可能过滤删除行之前统计行级计数
        try:
            added_count = int((merged_df['更新情况（标记）'] == '新增').sum())
            deleted_count = int((merged_df['更新情况（标记）'] == '删除').sum())
            updated_count = int((merged_df['更新情况（标记）'] == '更新').sum())
        except Exception:
            added_count = 0
            deleted_count = 0
            updated_count = 0
                
        # 删除比对过程中生成的临时列
        columns_to_drop = []
        for c in merged_df.columns:
            # 丢弃内部的 _ANCHOR 和 DIFF_ 列
            if c.startswith('_ANCHOR') or c.startswith('DIFF_'):
                columns_to_drop.append(c)
            # 丢弃 _OLD_ 后缀的列，除非它的原始列在 `deleted_cols` 中（表示是删除的列，其数据被拷贝回了主列）
            elif c.endswith('_OLD_') and c.replace('_OLD_', '') not in deleted_cols:
                columns_to_drop.append(c)
        
        # 当不合并删除数据时，所有 _OLD_ 列都应被删除
        if not getattr(config, 'merge_deleted_data', True):
            columns_to_drop = list(set(list(columns_to_drop) + [c for c in merged_df.columns if c.endswith('_OLD_')]))
        raw_output_df = merged_df.drop(columns=columns_to_drop, errors='ignore')
        # 再次确保"更新情况（标记）"列在第一列
        raw_output_df = reorder_columns_with_update_mark_first(raw_output_df)
        raw_output_df = raw_output_df.reset_index(drop=True) # 重置索引，使之连续

        # 当不合并删除数据时，剔除纯删除的记录
        if not getattr(config, 'merge_deleted_data', True) and '更新情况（标记）' in raw_output_df.columns:
            raw_output_df = raw_output_df[raw_output_df['更新情况（标记）'] != '删除']

        # 当不合并删除数据时，彻底删除旧有新无的列（deleted_cols）
        if not getattr(config, 'merge_deleted_data', True) and deleted_cols:
            raw_output_df = raw_output_df.drop(columns=[c for c in raw_output_df.columns if c in deleted_cols], errors='ignore')

        # 确定Sheet的总体变更类型
        change_type = 'data_changed' if has_changes or added_cols or deleted_cols else None
        
        # 设置结果DataFrame的SAS列名和列标签元数据
        sas_file_names = list(raw_output_df.columns) # 包含所有列名，包括标记列

        complete_sas_labels = [] # 存储完整的列标签
        for col_name in raw_output_df.columns:
            if col_name == '更新情况（标记）':
                complete_sas_labels.append('更新情况（标记）')
            else:
                # 从合并后的映射中获取标签，如果没有则使用列名本身
                complete_sas_labels.append(merged_name_to_label.get(col_name, col_name))
        sas_file_labels = complete_sas_labels # 包含所有列的标签，包括标记列的标签

        # 返回处理后的DataFrame和所有差异信息
        return raw_output_df, diff_dict, added_cols, deleted_cols, sas_file_names, sas_file_labels, merged_name_to_label, change_type, updated_count, deleted_count, added_count
        
    except Exception as e:
        # 针对常见的Pandas模糊布尔异常，给出更明确的中文提示
        msg = str(e)
        if 'The truth value of a Series is ambiguous' in msg:
            error_msg = (
                "锚点行可能存在重复列或无效值，导致行匹配失败。" 
                "请检查锚点行是否有重复列名，并确保锚点设置正确。\n"
                f"原始错误: {msg}"
            )
        else:
            error_msg = f"执行数据比对时出错: {msg}"
        progress_manager.safe_log(f"❌ {error_msg}") # 记录错误日志
        # 出错时返回空/默认值，确保程序不会崩溃
        return pd.DataFrame(), {}, [], [], [], [], {}, None, 0, 0, 0

def create_anchor_by_sas_names(df, key_sas_names, log_func, sheet_name=""):
    """
    基于指定的SASFieldName列表创建DataFrame的锚点列（_ANCHOR）。
    锚点列用于唯一标识每一行，是比对过程中行匹配的关键。
    Args:
        df (pd.DataFrame): 输入的数据DataFrame。
        key_sas_names (list): 作为锚点的SASFieldName（列名）列表。
        log_func (callable): 日志函数，用于记录处理过程中的信息。
        sheet_name (str, optional): 当前处理的表单名称，用于日志输出。默认为空字符串。
    Returns:
        pd.DataFrame: 添加了锚点列（_ANCHOR）的DataFrame。
                      如果无法创建有效锚点，_ANCHOR列可能为空或包含重复值。
    """
    # 获取DataFrame的SASFieldName信息，优先从attrs中获取，否则使用DataFrame的实际列名
    sas_names = getattr(df, 'attrs', {}).get('sas_file_name', [])
    if not sas_names:
        # 如果没有SASFieldName信息，无法创建有效锚点，锚点列置空
        log(f"    ⚠️ 表单[{sheet_name}]没有SASFieldName信息，锚点列置空", log_func)
        df['_ANCHOR'] = ""
        return df

    # 查找所有在当前DataFrame中存在的锚点列
    sas_names_set = set(sas_names)
    matched_keys = [key for key in key_sas_names if key in sas_names_set]
   
    if not matched_keys:
        # 如果没有找到任何匹配的锚点列，锚点列置空
        log(f"    ⚠️ 表单[{sheet_name}]未找到任何匹配的锚点，锚点列置空", log_func)
        df['_ANCHOR'] = ""
        return df
    
    # 获取SAS列名到标签的映射（用于日志或调试，尽管此处主要使用列名）
    try:
        sas_name_to_label = getattr(df, 'attrs', {}).get('sas_name_to_label', {}) 
    except:
        log(f"    ⚠️ 表单[{sheet_name}]获取SASFieldName和SASFieldLabel的映射关系失败，锚点列置空", log_func)
        df['_ANCHOR'] = ""
        return df
    
    # 如果锚点列与实际DataFrame列不一致，直接置空，避免后续KeyError
    missing_keys = [key for key in matched_keys if key not in df.columns]
    if missing_keys:
        log(f"    ⚠️ 表单[{sheet_name}]锚点列缺失: {', '.join(missing_keys)}，锚点列置空", log_func)
        df['_ANCHOR'] = ""
        return df
    
    try:
        # 添加锚点列(_ANCHOR)，将所有关键列的值用"###"连接起来作为唯一标识
        # 确保所有值都被转换为字符串，避免类型不一致导致的问题
        df['_ANCHOR'] = df[matched_keys].astype(str).agg('###'.join, axis=1)

        # 检查锚点列是否有重复值，如果存在则记录警告
        if not df['_ANCHOR'].is_unique:
            log(f"    ⚠️ 锚点重复，重复数量: {df['_ANCHOR'].duplicated().sum()}", log_func)

    except Exception as e:
        log(f"    ⚠️ 创建锚点时出错: {str(e)}", log_func)
        # 确保_ANCHOR列存在，即使出错也将其初始化为空字符串
        if '_ANCHOR' not in df.columns:
            df['_ANCHOR'] = ''
    
    return df

def compare_columns_by_sas_names(merged_df, key_sas_names, sheet_name, log_func):
    """
    基于SASFieldName比对合并后的DataFrame的列差异。
    此函数会找出单元格级别的新增、删除和更新，并生成一个差异记录DataFrame。
    Args:
        merged_df (pd.DataFrame): 包含新旧版本数据合并后的DataFrame。
                                  其中旧版本列名带有"_OLD_"后缀。
        sheet_name (str): 当前处理的表单名称。
        key_sas_names (list): 锚点列的SASFieldName列表，这些列将不参与比对。
        log_func (callable): 日志函数。
    Returns:
        pd.DataFrame: 差异记录DataFrame，包含'row'（行索引）、'col'（列名）和'flag'（差异类别：'新增'/'删除'/'更新'）
                      如果无差异，则返回空DataFrame。
    """
    dift = pd.DataFrame() # 初始化空的差异记录DataFrame

    # 获取合并表单的所有列名
    all_columns = list(merged_df.columns)
    
    # 找到所有带有"_OLD_"后缀的旧版本列
    old_columns = [col for col in all_columns if col.endswith("_OLD_")]
    
    # 找到新旧版本共有的列（即同时存在原列名和带"_OLD_"后缀的列名）
    compare_sas_names = []
    for old_col in old_columns:
        original_col = old_col.replace("_OLD_", "") # 去掉"_OLD_"后缀得到原列名
        if original_col in all_columns:
            compare_sas_names.append(original_col)
        
    # 过滤掉锚点列，只保留需要比对的非锚点列
    # 排除_ANCHOR和更新情况（标记）列，因为它们是辅助列
    non_anchor_compare_cols = [col for col in compare_sas_names if col not in key_sas_names and col != '_ANCHOR' and col != '更新情况（标记）']

    diff_records = [] # 存储差异记录的列表，后续转换为DataFrame
    diff_columns = set() # 存储差异标记列的名称（用于清理）
    
    # 验证diff_records初始化状态（防御性编程，以防NoneType错误）
    if diff_records is None:
        diff_records = []
        log_func(f"警告: diff_records 初始化为None，已重新初始化为空列表")
    
    # 优化：预计算所有需要比较的列的空值状态和字符串值，减少循环内的重复计算
    null_status_cache = {}
        
    def is_null_optimized(value):
        """优化的空值检查函数：检查值是否为None、Pandas的NaN或只包含空白字符的字符串。"""
        if pd.isna(value):
            return True
        if isinstance(value, str):
            return value.strip() == ''
        return False

    def is_numeric_value(value):
        if pd.isna(value):
            return False
        return isinstance(value, Number) and not isinstance(value, bool)
    
    # 预处理：批量检查空值状态，并将列转换为字符串类型，提高性能
    for col in non_anchor_compare_cols:           
        old_col = f"{col}_OLD_"
        if col not in merged_df.columns or old_col not in merged_df.columns: # 确保新旧列都存在
            continue
            
        col_new = merged_df[col]
        col_old = merged_df[old_col]
        
        new_text = col_new.astype(str).str.strip()
        old_text = col_old.astype(str).str.strip()
        new_null_mask = col_new.isna() | (new_text == '')
        old_null_mask = col_old.isna() | (old_text == '')
        new_is_num = col_new.map(is_numeric_value)
        old_is_num = col_old.map(is_numeric_value)
        new_text_num = pd.to_numeric(new_text, errors='coerce')
        old_text_num = pd.to_numeric(old_text, errors='coerce')
        
        null_status_cache[col] = {
            'new_null': new_null_mask,
            'old_null': old_null_mask,
            'new_text': new_text,
            'old_text': old_text,
            'new_is_num': new_is_num,
            'old_is_num': old_is_num,
            'new_text_num': new_text_num,
            'old_text_num': old_text_num
        }
    
    # 遍历所有需要比较的非锚点列
    check_counter = [0]  # 用于计数的列表，用于控制停止检查的频率
    for col in non_anchor_compare_cols:
        # 提高检查频率：每20次列处理检查一次
        if check_counter[0] % 20 == 0:
            check_stop_frequently(log_func)
        check_counter[0] += 1
        
        if col not in merged_df.columns or f"{col}_OLD_" not in merged_df.columns:
            log_func(f"    ⚠️ 表单[{sheet_name}]列[{col}]在合并数据中丢失，跳过比较")
            continue # 列缺失，跳过比对
        
        diff_col_name = f"DIFF_{col}" # 差异标记列名（例如DIFF_FIELD_NAME）
        diff_columns.add(diff_col_name)
        
        # 从缓存中获取预处理的数据，避免重复计算
        col_new = merged_df[col]
        col_old = merged_df[f"{col}_OLD_"]

        if col in null_status_cache:
            cache_data = null_status_cache[col]
            new_null_mask = cache_data['new_null']
            old_null_mask = cache_data['old_null']
            new_text = cache_data['new_text']
            old_text = cache_data['old_text']
            new_is_num = cache_data['new_is_num']
            old_is_num = cache_data['old_is_num']
            new_text_num = cache_data['new_text_num']
            old_text_num = cache_data['old_text_num']
        else:
            # 如果不在缓存中（理论上不应该发生），回退到原始方法
            col_new = merged_df[col]
            col_old = merged_df[f"{col}_OLD_"]
            new_text = col_new.astype(str).str.strip()
            old_text = col_old.astype(str).str.strip()
            new_null_mask = col_new.isna() | (new_text == '')
            old_null_mask = col_old.isna() | (old_text == '')
            new_is_num = col_new.map(is_numeric_value)
            old_is_num = col_old.map(is_numeric_value)
            new_text_num = pd.to_numeric(new_text, errors='coerce')
            old_text_num = pd.to_numeric(old_text, errors='coerce')
        
        # 创建差异标记列，初始化为False
        merged_df[diff_col_name] = False
        
        # 使用向量化操作进行批量比较，大幅提升性能
        # 找出新增的单元格：新数据中有值，旧数据中无值
        new_addition_mask = (~new_null_mask) & old_null_mask
        # 找出删除的单元格：新数据中无值，旧数据中有值  
        deletion_mask = new_null_mask & (~old_null_mask)
        # 找出更新的单元格：新旧数据都有值，但值不相同
        non_null_mask = (~new_null_mask) & (~old_null_mask)
        mask_num_num = new_is_num & old_is_num & non_null_mask
        mask_text_text = (~new_is_num) & (~old_is_num) & non_null_mask
        mask_new_text_old_num = (~new_is_num) & old_is_num & non_null_mask
        mask_new_num_old_text = new_is_num & (~old_is_num) & non_null_mask

        diff_num_num = mask_num_num & (col_new != col_old)
        diff_text_text = mask_text_text & (new_text != old_text)

        new_text_num_ok = new_text_num.notna()
        old_text_num_ok = old_text_num.notna()

        diff_new_text_old_num = mask_new_text_old_num & (
            (new_text_num_ok & (new_text_num != col_old)) |
            (~new_text_num_ok & (new_text != old_text))
        )
        diff_new_num_old_text = mask_new_num_old_text & (
            (old_text_num_ok & (col_new != old_text_num)) |
            (~old_text_num_ok & (new_text != old_text))
        )

        update_mask = diff_num_num | diff_text_text | diff_new_text_old_num | diff_new_num_old_text
        
        # 批量标记差异列为True
        merged_df.loc[new_addition_mask, diff_col_name] = True
        merged_df.loc[deletion_mask, diff_col_name] = True  
        merged_df.loc[update_mask, diff_col_name] = True
        
        # 批量创建差异记录列表
        # 记录新增
        if new_addition_mask.any(): # 如果存在新增的行
            addition_indices = new_addition_mask[new_addition_mask].index # 获取新增行的索引
            for idx in addition_indices:
                try:
                    if diff_records is None: # 防御性编程
                        diff_records = []
                        log_func(f"警告: diff_records 初始化为None，已重新初始化为空列表")
                    
                    diff_records.append({
                        'row': idx, 
                        'col': col,
                        'flag': '新增' # 标记为新增
                    })
                except AttributeError as e:
                    log_func(f"错误: diff_records.append失败: {str(e)}")
                    # 重新初始化并重试，处理极端情况
                    diff_records = []
                    diff_records.append({
                        'row': idx, 
                        'col': col,
                        'flag': '新增'
                    })
                except Exception as e:
                    log_func(f"差异记录append时出现未知错误: {str(e)}")
        
        # 记录删除
        if deletion_mask.any(): # 如果存在删除的行
            deletion_indices = deletion_mask[deletion_mask].index # 获取删除行的索引
            for idx in deletion_indices:
                try:
                    if diff_records is None: # 防御性编程
                        diff_records = []
                        log_func(f"警告: diff_records在删除记录append时为None，已重新初始化")
                    
                    diff_records.append({
                        'row': idx, 
                        'col': col,
                        'flag': '删除' # 标记为删除
                    })
                except Exception as e:
                    log_func(f"删除记录append时出错: {str(e)}")
                    if diff_records is None:
                        diff_records = []
                        diff_records.append({
                            'row': idx, 
                            'col': col, 
                            'flag': '删除'
                        })
                
        # 记录更新
        if update_mask.any(): # 如果存在更新的行
            update_indices = update_mask[update_mask].index # 获取更新行的索引
            for idx in update_indices:
                try:
                    if diff_records is None: # 防御性编程
                        diff_records = []
                        log_func(f"警告: diff_records在更新记录append时为None，已重新初始化")
                    
                    diff_records.append({
                        'row': idx, 
                        'col': col, 
                        'flag': '更新' # 标记为更新
                    })
                except Exception as e:
                    log_func(f"更新记录append时出错: {str(e)}")
                    if diff_records is None:
                        diff_records = []
                        diff_records.append({
                            'row': idx, 
                            'col': col, 
                            'flag': '更新'
                        })
    
    # 创建差异记录DataFrame
    try:
        if diff_records and len(diff_records) > 0:
            dift = pd.DataFrame(diff_records) # 将记录列表转换为DataFrame
            # print(f"✅ Sheet [{sheet_name}] 发现 {len(dift)} 条差异记录。首几条：\n{dift.head().to_string()}")
        else:
            dift = pd.DataFrame(columns=['row', 'col', 'flag']) # 如果没有差异，返回空的DataFrame
            log_func(f"   ℹ️ Sheet [{sheet_name}] 未发现差异记录。")
    except Exception as e:
        log_func(f"❌ 创建差异记录DataFrame时出错: {str(e)}")
        dift = pd.DataFrame(columns=['row', 'col', 'flag']) # 确保在出错时也返回空DataFrame结构

    return dift

def process_edc_multithreaded(old_path, new_path, output_path, log_func, config=None, progress_func=None, stop_flag=None):
    """
    多线程优化版的EDC处理主函数。
    负责协调整个比对流程：文件预处理、多线程Sheet比对、结果合并和文件保存。
    支持大表单数据拆分处理，并通过线程安全机制确保稳定性和性能。
    
    Args:
        old_path (str): 旧版本Excel文件路径。
        new_path (str): 新版本Excel文件路径。
        output_path (str): 结果Excel文件的输出路径。
        log_func (callable): 日志记录函数。
        config (ConfigManager, optional): 配置对象。如果为None，则使用全局配置变量。
        progress_func (callable, optional): GUI进度更新回调函数。
        stop_flag (threading.Event, optional): 停止标志，用于控制线程中断。默认为None。
    Returns:
        str: 最终输出的Excel文件路径。
    Raises:
        RuntimeError: 如果处理过程中发生致命错误（例如文件保存失败）。
        InterruptedError: 如果用户请求停止操作。
    """
    set_global_stop_flag(stop_flag) # 设置全局停止标志，以便所有线程都能访问
    
    # 在函数开始处创建最终结果工作簿
    final_wb = Workbook()
    final_wb.remove(final_wb.active) # 删除默认创建的空Sheet
    has_any_sheet_data = False # 标记是否有任何 Sheet 成功写入数据

    # 获取配置，优先使用传入的config对象，否则使用全局变量
    # 这是一个兼容旧版全局变量和新版ConfigManager的方式
    # EDC = config.edc if config else EDC # 这行需要从原始文件中删除，因为EDC不再是全局变量
    # EXCLUDE_SHEETS = config.exclude_sheets if config else EXCLUDE_SHEETS # 同上
    # COMMON_COLS_TO_DROP = config.common_cols_to_drop if config else COMMON_COLS_TO_DROP # 同上
    # DEFAULT_KEYS = config.default_keys if config else DEFAULT_KEYS # 同上
    # SHEET_KEY_MAP = config.sheet_key_map if config else SHEET_KEY_MAP # 同上
    # HIGHLIGHT_FILL = config.highlight_fill if config else HIGHLIGHT_FILL # 同上
    # MISSING_SHEET_TAB_FILL = config.missing_sheet_tab_fill if config else MISSING_SHEET_TAB_FILL # 同上
    # NEW_SHEET_TAB_FILL = config.new_sheet_tab_fill if config else NEW_SHEET_TAB_FILL # 同上

    # 使用传入的config对象来获取配置
    anchor_row_num = config.anchor_row_num
    header_row_num = config.header_row_num
    exclude_sheets = config.exclude_sheets
    common_cols_to_drop = config.common_cols_to_drop
    default_keys = config.default_keys
    sheet_key_map = config.sheet_key_map
    highlight_fill = config.highlight_fill
    missing_sheet_tab_fill = config.missing_sheet_tab_fill
    new_sheet_tab_fill = config.new_sheet_tab_fill
    
    try:
        check_stop_frequently(log_func, stop_flag)  # 在开始时检查是否需要停止
        
        # 文件预处理：解除保护、清除筛选器、删除排除的Sheet
        log_func("正在预处理旧版本文件...")
        old_result = check_and_remove_file_protection(old_path, exclude_sheets, log_func)
        current_old_path = old_result[2] if old_result and len(old_result) > 2 else old_path # 获取处理后的旧文件路径
        
        log_func("正在预处理新版本文件...")
        new_result = check_and_remove_file_protection(new_path, exclude_sheets, log_func)
        current_new_path = new_result[2] if new_result and len(new_result) > 2 else new_path # 获取处理后的新文件路径
        
        log_func("开始获取所有表单名称...")
      
        # 获取所有Sheet名称，不加载数据，提高效率
        old_sheet_names = get_sheet_names(current_old_path, log_func)
        new_sheet_names = get_sheet_names(current_new_path, log_func)

        # 合并新旧文件中的所有Sheet名称，并去除重复，找出需要处理的Sheet
        all_sheets = set()
        all_sheets.update(old_sheet_names)
        all_sheets.update(new_sheet_names)

        # 记录将被排除的表单（在此阶段就会被过滤，因此后续函数内不会再触发"跳过"日志）
        excluded_sheets_found = sorted([s for s in all_sheets if s in exclude_sheets])
        if excluded_sheets_found:
            log(f"⏭️ 跳过 {len(excluded_sheets_found)} 个表单（在排除列表中）: {', '.join(excluded_sheets_found)}", log_func)
        # 记录配置中存在但文件中不存在的排除项，帮助排查配置问题
        unmatched_excludes = sorted([s for s in exclude_sheets if s not in all_sheets])
        if unmatched_excludes:
            log(f"ℹ️ 排除列表中存在文件中未找到的表单: {', '.join(unmatched_excludes)}", log_func)

        sheets_to_process = [sheet for sheet in all_sheets if sheet not in exclude_sheets] # 排除不需要处理的Sheet
        sheets_to_process.sort() # 确保处理顺序一致，方便调试和跟踪

        log_func(f"共发现 {len(sheets_to_process)} 个表单需要处理")

        if not sheets_to_process:
            log_func("没有需要处理的表单，程序结束。")
            update_progress("完成", 100, progress_func=progress_func) # 更新最终进度
            # 如果没有需要处理的Sheet，创建一个空的工作簿并保存，表示无差异或跳过所有
            final_wb.remove(final_wb.active) # 删除默认创建的空Sheet
            # 插入汇总表
            summary_ws = final_wb.create_sheet(title="比对结果汇总", index=0)
            summary_ws.append(["Sheet 名称", "更新行数", "删除行数", "新增行数"]) 
            # 插入说明表
            empty_sheet = final_wb.create_sheet("无差异表单") # 创建一个说明Sheet
            empty_sheet.append(["说明"])
            empty_sheet.append(["所有表单均无差异或处理失败"])
            # 对所有表应用基础样式并保存
            enable_filter_border_header_all_sheets(final_wb, output_path)
            return output_path # 返回输出路径

        # 设置最大并发工作线程数：优先使用配置中的值，否则采用CPU核心数-1
        configured_workers = None
        try:
            configured_workers = getattr(config, 'max_workers', None)
        except Exception:
            configured_workers = None
        if not isinstance(configured_workers, int) or configured_workers <= 0:
            configured_workers = max(1, (os.cpu_count() or 1) - 1)
        max_workers = min(len(sheets_to_process), configured_workers)
        
        # 初始化线程安全的进度管理器
        progress_manager = ThreadSafeProgressManager(len(sheets_to_process), progress_func, log_func)
        
        # 移除 sheet_results 列表
        # sheet_results = [] # 存储每个Sheet的处理结果
        
        update_progress("开始处理表单...", 20, progress_func=progress_func) # 更新GUI进度

        # 使用线程池执行Sheet处理任务
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            # 提交所有Sheet处理任务
            future_to_sheet = {
                executor.submit(
                    process_single_sheet_complete,
                    sheet_name,
                    current_old_path,
                    current_new_path,
                    config,
                    progress_manager,
                    # global_split_config # Removed
                ): sheet_name
                for sheet_name in sheets_to_process
            }
            
            # 遍历已完成的任务，收集结果并直接写入（支持停止）
            stop_requested = False
            for future in concurrent.futures.as_completed(future_to_sheet):
                # 如果用户请求停止，尽快跳出并取消未开始的任务
                if stop_flag and stop_flag.is_set():
                    stop_requested = True
                    break
                sheet_name = future_to_sheet[future]
                try:
                    result = future.result() # 获取任务结果
                    
                    if not result.success:
                        log_func(f"⚠️ 表单 [{sheet_name}] 处理失败: {result.error_message}") # 记录失败信息
                        continue # 跳过当前失败的 Sheet
                    
                    # 检查是否明确因为新旧数据都为空而跳过比对（仅有表头）
                    # change_type 为 None 表示没有实际数据变更，diffs 为空表示没有单元格差异
                    if result.change_type is None and result.df is not None and result.df.empty and not result.differences:
                        print(f"   ⏩ sheet[{result.sheet_name}]新旧版本均为空且无差异，跳过合并。)") # 调整消息，更清晰
                        continue # 跳过不写入
                    
                    # 如果DataFrame对象为None，则跳过
                    if result.df is None: 
                        log_func(f"⚠️ Sheet [{result.sheet_name}] 没有处理结果DataFrame，跳过合并。")
                        continue # 如果DataFrame是None，跳过当前Sheet

                    # 在最终工作簿中创建Sheet
                    target_ws = final_wb.create_sheet(title=result.sheet_name) 
                    has_any_sheet_data = True # 标记有数据写入

                    # 写入表头，优先使用结果中的SAS列标签，否则使用DataFrame的列名
                    if result.sas_file_labels:
                        target_ws.append(result.sas_file_labels) # 写入标签作为表头
                    else:
                        target_ws.append(result.df.columns.tolist()) # 回退到DataFrame列名

                    # 仅当DataFrame不为空时才写入数据行
                    if not result.df.empty: 
                        for r_idx, r in enumerate(dataframe_to_rows(result.df, index=False, header=False)):
                            target_ws.append(r)
                            if r_idx % 5000 == 0: # 每5000行执行一次垃圾回收，降低内存峰值
                                # 将 gc.collect() 替换为条件垃圾回收
                                _maybe_gc_collect(threshold_percent=70, log_func=log_func)
                                # 同步检查是否需要停止
                                check_stop_frequently(log_func, stop_flag)
                    else:
                        # 为空数据且仅有表头的表单添加日志
                        log_func(f"ℹ️ Sheet [{result.sheet_name}] 数据为空，仅写入表头。")

                    # 根据比对结果应用高亮显示
                    apply_highlight_to_worksheet(
                        ws=target_ws,
                        config=config,
                        sheet_type=result.change_type,
                        diff_info=result.differences,
                        add_sas_names=result.add_sas_names,
                        del_sas_names=result.del_sas_names,
                        sas_file_names=result.sas_file_names,
                        log_func=log_func
                    )
                    
                    # 更新SAS表头信息和Sheet标签颜色（这些信息从SheetProcessResult传递）
                    target_ws._add_sas_names = result.add_sas_names
                    target_ws._del_sas_names = result.del_sas_names
                    target_ws._sas_file_names = result.sas_file_names
                    target_ws._sas_file_labels = result.sas_file_labels
                    target_ws._sas_name_to_label = result.sas_name_to_label # 确保列名到标签的映射被传递
                    # 记录行级计数，供汇总表使用
                    try:
                        target_ws._updated_rows_count = int(getattr(result, 'updated_rows_count', 0) or 0)
                        target_ws._deleted_rows_count = int(getattr(result, 'deleted_rows_count', 0) or 0)
                        target_ws._added_rows_count = int(getattr(result, 'added_rows_count', 0) or 0)
                    except Exception:
                        target_ws._updated_rows_count = 0
                        target_ws._deleted_rows_count = 0
                        target_ws._added_rows_count = 0

                    
                    # 应用Sheet标签颜色
                    if result.change_type == 'missing':
                        target_ws.sheet_properties.tabColor = config.missing_sheet_tab_fill.start_color # 缺失Sheet颜色
                    elif result.change_type == 'new':
                        target_ws.sheet_properties.tabColor = config.new_sheet_tab_fill.start_color # 新增Sheet颜色
                    elif result.change_type == 'data_changed':
                        target_ws.sheet_properties.tabColor = config.highlight_fill.start_color # 数据有变化Sheet颜色
                    
                    print(f"✅ sheet[ {result.sheet_name}]已完成写入")
                        
                except InterruptedError:
                    # 显式传播中断，以便外层捕获并终止
                    log_func(f"表单 [{sheet_name}] 处理被用户停止")
                    # 取消尚未完成的任务
                    for f in future_to_sheet.keys():
                        if not f.done():
                            f.cancel()
                    raise
                except Exception as exc:
                    error_msg = f"表单 [{sheet_name}] 写入最终文件时异常: {str(exc)}"
                    log_func(error_msg) # 记录异常信息
                    # 不再将失败结果添加到 sheet_results
            
            if stop_requested:
                # 取消未完成的任务，并中断流程
                for f in future_to_sheet.keys():
                    if not f.done():
                        f.cancel()
                raise InterruptedError("用户停止了操作")

    except InterruptedError:
        log_func("操作已被用户停止，终止多线程执行。")
        raise # 重新抛出到主线程，以便GUI可以捕获并显示停止信息
    except Exception as executor_error:
        error_msg = f"多线程执行出错: {str(executor_error)}\n"
        log_func(error_msg)
        raise RuntimeError(error_msg) from executor_error # 抛出运行时错误，表示致命问题
    
    update_progress("所有表单处理完成，正在保存最终文件...", 85, progress_func=progress_func) # 更新GUI进度
    log_func("所有表单处理完成，开始保存最终文件...")

    # 如果没有成功处理的Sheet（即final_wb仍然是空的，只有默认创建后又删除的那个），则创建说明Sheet
    if not has_any_sheet_data and not final_wb.sheetnames:
        empty_sheet = final_wb.create_sheet("无差异表单")
        empty_sheet.append(["说明"])
        empty_sheet.append(["所有表单均无差异或处理失败"])
        log_func("创建空结果表单")
    
    # 在保存前插入"比对结果汇总"工作表到第一个位置
    try:
        summary_ws = final_wb.create_sheet(title="比对结果汇总", index=0)
        # 表头
        summary_ws.append(["Sheet 名称", "更新行数", "删除行数", "新增行数"])
        # 遍历后续工作表，生成汇总（仅统计存在差异的表单）
        for ws in final_wb.worksheets[1:]:
            try:
                sheet_name = ws.title
                # 优先使用写入阶段存储的计数
                updated_cnt = int(getattr(ws, '_updated_rows_count', 0) or 0)
                deleted_cnt = int(getattr(ws, '_deleted_rows_count', 0) or 0)
                added_cnt = int(getattr(ws, '_added_rows_count', 0) or 0)
                # 若计数全为0，再回退扫描首列标记计数
                if (updated_cnt + deleted_cnt + added_cnt) == 0:
                    header_map = {cell.value: idx+1 for idx, cell in enumerate(ws[1]) if cell.value}
                    update_mark_col_idx = header_map.get('更新情况（标记）', None)
                    if update_mark_col_idx is not None:
                        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=update_mark_col_idx, max_col=update_mark_col_idx):
                            cell = row[0]
                            v = str(cell.value).strip() if cell.value is not None else ''
                            if v == '更新':
                                updated_cnt += 1
                            elif v == '删除':
                                deleted_cnt += 1
                            elif v == '新增':
                                added_cnt += 1
                # 仅当存在差异（任一计数>0）才写入汇总
                if (updated_cnt + deleted_cnt + added_cnt) > 0:
                    row_idx = summary_ws.max_row + 1
                    # 写名称
                    name_cell = summary_ws.cell(row=row_idx, column=1, value=sheet_name)
                    # 设置指向对应sheet A1 的超链接
                    safe_sheet = sheet_name.replace("'", "''")
                    name_cell.hyperlink = f"#'{safe_sheet}'!A1"
                    name_cell.style = "Hyperlink"
                    # 写计数
                    summary_ws.cell(row=row_idx, column=2, value=updated_cnt)
                    summary_ws.cell(row=row_idx, column=3, value=deleted_cnt)
                    summary_ws.cell(row=row_idx, column=4, value=added_cnt)
            except Exception as e:
                log_func(f"生成汇总时处理表单[{ws.title}]出错: {str(e)}")
    except Exception as e:
        log_func(f"创建'比对结果汇总'工作表失败: {str(e)}")
    
    check_stop_frequently(log_func, stop_flag) # 在保存前再次检查是否停止
    
    update_progress("正在保存文件...", 90, progress_func=progress_func) # 更新GUI进度
    log_func("正在保存文件...")

    try:
        enable_filter_border_header_all_sheets(final_wb, output_path)
        # final_wb.save(output_path) # 保存最终结果Excel文件
        update_progress("文件保存完成", 100, progress_func=progress_func) # 更新最终进度
        log_func("✅ 文件保存完成")
    except Exception as save_error:
        error_msg = f"保存文件失败: {str(save_error)}"
        log_func(error_msg)
        raise RuntimeError(error_msg) from save_error # 保存失败是致命错误，抛出异常
    finally:
        # 确保工作簿关闭，释放资源
        try:
            final_wb.close()
        except Exception as e:
            log_func(f"⚠️ 关闭工作簿时出错: {str(e)}")
        # 清理所有 nofilter 缓存文件（静默，不写入日志）
        try:
            cleanup_nofilter_files()
        except Exception:
            pass
    
    # 清理预处理阶段生成的临时文件和参数文件（静默，不写入日志）
    try:
        for temp_path in [current_old_path, current_new_path]: 
            if temp_path and temp_path.endswith('_nofilter.xlsx') and os.path.exists(temp_path):
                os.remove(temp_path)
    except Exception:
        pass
    
    try:
        app_temp_dir = get_app_temp_dir()
        parameters_file_path = os.path.join(app_temp_dir, 'parameters.json')
        if os.path.exists(parameters_file_path):
            os.remove(parameters_file_path)
    except Exception:
        pass
    
    return output_path 


def enable_filter_border_header_all_sheets(wb, output_path):
    thin_side = Side(border_style="thin", color="000000")  # 细黑边框
    border = Border(
        left=thin_side,
        right=thin_side,
        top=thin_side,
        bottom=thin_side
    )
    bold_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True)
    # 默认表头填充色（与高亮工具中保持一致）
    default_header_fill = PatternFill(start_color='A6C9EC', end_color='A6C9EC', fill_type='solid')

    for ws in wb.worksheets:
        max_row = ws.max_row
        max_col = ws.max_column

        if max_row > 1 and max_col > 0:
            # 设置筛选范围
            filter_range = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.auto_filter.ref = filter_range

            # 全区域加边框
            for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
                for cell in row:
                    cell.border = border
            ws.freeze_panes = "A2"
            # 首行字体加粗 + 自动换行
            for cell in ws[1]:
                cell.font = bold_font
                cell.alignment = wrap_alignment
                # 如果表头单元格尚未设置填充，则应用默认表头填充色
                try:
                    if not getattr(cell.fill, 'fill_type', None):
                        cell.fill = default_header_fill
                except Exception:
                    # 避免任何由于单元格状态导致的异常阻断保存
                    pass

    wb.save(output_path)
