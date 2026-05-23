import pandas as pd
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# 从其他模块导入必要的类和函数
from ..utils.config_manager import ConfigManager # 导入ConfigManager

class HighlightOptimizer:
    """高亮标记优化器，提供缓存和批量处理功能。
    旨在提高Excel写入时高亮操作的性能，通过缓存单元格值和行状态来减少重复计算和Openpyxl的API调用。
    """
    
    def __init__(self):
        """初始化高亮优化器。"""
        # 缓存字典
        self.cell_value_cache = {}  # 缓存单元格值，格式为 {(row_idx, col_idx): value_str}
        self.row_status_cache = {}  # 缓存行状态，格式为 {(sheet_title, row_idx): status_dict}
        self.empty_rows_cache = set()  # 缓存空行信息（行索引集合）
        self.non_empty_rows_cache = set()  # 缓存非空行信息（行索引集合）
        
    def clear_cache(self):
        """清空所有缓存，在处理新的工作表或文件前调用。"""
        self.cell_value_cache.clear()
        self.row_status_cache.clear()
        self.empty_rows_cache.clear()
        self.non_empty_rows_cache.clear()
    
    def get_cell_value_cached(self, cell):
        """获取单元格值（带缓存）。
        Args:
            cell (openpyxl.cell.cell.Cell): Openpyxl单元格对象。
        Returns:
            str: 单元格的字符串值，去除首尾空格。
        """
        cell_key = (cell.row, cell.column) # 使用行号和列号作为缓存键
        if cell_key not in self.cell_value_cache:
            value = cell.value
            self.cell_value_cache[cell_key] = str(value).strip() if value else "" # 缓存处理后的字符串值
        return self.cell_value_cache[cell_key]
    
    def is_row_empty_cached(self, sheet, row_idx, start_col=1, end_col=None):
        """检查行是否为空（带缓存）。
        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet): Openpyxl工作表对象。
            row_idx (int): 行索引（1-based）。
            start_col (int, optional): 开始检查的列索引（1-based）。默认为1。
            end_col (int, optional): 结束检查的列索引（1-based）。默认为工作表的最后一列。
        Returns:
            bool: 如果行为空则返回True。
        """
        if row_idx in self.empty_rows_cache:
            return True # 已知为空行，直接返回
        if row_idx in self.non_empty_rows_cache:
            return False # 已知为非空行，直接返回
            
        # 实际检查逻辑
        if end_col is None:
            end_col = sheet.max_column
            
        is_empty = True
        row = sheet[row_idx] # 获取整行数据
        for col_idx in range(start_col - 1, min(end_col, len(row))): # 遍历指定列范围
            if col_idx < len(row):
                cell = row[col_idx]
                if cell and cell.value is not None and str(cell.value).strip():
                    is_empty = False # 发现非空单元格
                    break
        
        # 缓存结果
        if is_empty:
            self.empty_rows_cache.add(row_idx)
        else:
            self.non_empty_rows_cache.add(row_idx)
            
        return is_empty
    
    def batch_get_category_values(self, sheet, category_col_letter, start_row, end_row):
        """批量获取"更新情况（标记）"列的值。
        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet): Openpyxl工作表对象。
            category_col_letter (str): "更新情况（标记）"列的字母表示（如'A'）。
            start_row (int): 开始行索引（1-based）。
            end_row (int): 结束行索引（1-based）。
        Returns:
            dict: 格式为 {row_idx: category_value} 的字典。
        """
        category_values = {}
        try:
            category_column = sheet[category_col_letter] # 获取整列数据
            for row_idx in range(start_row, min(end_row + 1, len(category_column) + 1)): # 遍历指定行范围
                if row_idx - 1 < len(category_column):
                    cell = category_column[row_idx - 1]
                    if cell:
                        value = self.get_cell_value_cached(cell) # 使用缓存获取单元格值
                        category_values[row_idx] = value
        except Exception as e:
            # 如果批量获取失败，回退到逐个获取（此分支通常不应发生，除非Openpyxl内部错误）
            pass
        return category_values
    
    def get_row_status_cached(self, sheet, row_idx, category_value, diff_info, col_name_to_idx):
        """获取行状态（带缓存）。
        判断行是否需要高亮，并找出需要高亮的具体单元格。
        Args:
            sheet (openpyxl.worksheet.worksheet.Worksheet): Openpyxl工作表对象。
            row_idx (int): 行索引（1-based）。
            category_value (str): "更新情况（标记）"列的值（如"新增", "删除", "更新"）。
            diff_info (dict): 单元格级别的差异信息（来自compare_columns_by_sas_names的diff_dict）。
            col_name_to_idx (dict): 列名到列索引（1-based）的映射。
        Returns:
            dict: 包含行状态信息的字典。
        """
        cache_key = (sheet.title, row_idx) # 使用工作表标题和行号作为缓存键
        if cache_key in self.row_status_cache:
            return self.row_status_cache[cache_key] # 已知行状态，直接返回
        
        # 计算行状态
        status = {
            'category': category_value, # 变更类别
            'needs_highlight': False, # 是否需要高亮整行或部分单元格
            'highlight_cells': [], # 需要高亮的单元格列表 (存储列名或列索引，而非单元格对象)
            'is_empty': False # 行是否为空（目前未使用，保留）
        }
        
        if category_value in ["新增", "删除", "更新"]:
            status['needs_highlight'] = True # 标记该行需要高亮
            
            if category_value in ["新增", "删除"]:
                # 对于新增或删除的行，标记高亮整行 (通过设置highlight_cells为特殊的'ALL'标记)
                status['highlight_cells'] = ['ALL'] 
                    
            elif category_value == "更新":
                # 对于更新的行，返回需要高亮的列名列表
                data_row_idx = row_idx - 2 # df_row_idx是DataFrame的0-based索引，而Excel的row_idx是1-based，且包含表头。
                if diff_info and data_row_idx in diff_info: # 检查该行是否有详细差异信息
                    status['highlight_cells'] = list(diff_info[data_row_idx].keys()) # 存储有差异的列名
        
        # 缓存结果
        self.row_status_cache[cache_key] = status
        return status

# Global optimizer instance
highlight_optimizer = HighlightOptimizer()

def apply_highlight_to_worksheet(ws, config, sheet_type=None, diff_info=None, add_sas_names=[], del_sas_names=[], sas_file_names=[], log_func=None):
    """根据比对结果为工作表（包括表头、单元格和Sheet标签）应用高亮和颜色标记。
    "更新"的行只高亮有变化的单元格，"新增"或"删除"的行则高亮整行，并相应标记列的增删。
    
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Openpyxl工作表对象，比对结果将应用到此工作表。
        config (ConfigManager): 配置管理器对象，包含颜色等设置。
        sheet_type (str, optional): Sheet的变更类型（'new', 'missing', 'data_changed'）。
        diff_info (dict, optional): 单元格级别的差异信息字典，格式为 {df_row_idx: {col_name: flag}}。
                                    这里的df_row_idx是DataFrame的0-based索引。
        add_sas_names (list, optional): 新增的SAS列名列表。
        del_sas_names (list, optional): 删除的SAS列名列表。
        sas_file_names (list, optional): 最终结果DataFrame的SAS列名列表（即表头行显示的列名）。
        log_func (callable, optional): 日志函数，用于记录操作信息。
    """

    # 1. 表头染色 (Header highlighting)
    # 默认表头填充色（蓝色），用于无特定增删标记的表头单元格
    header_fill = PatternFill(start_color='A6C9EC', end_color='A6C9EC', fill_type='solid')
    # 获取工作表的第一行（表头行），如果工作表为空则返回空列表
    first_row = ws[1] if ws.max_row > 0 else []

    # 遍历第一行的每个单元格（即表头单元格）
    for col_idx, cell in enumerate(first_row,1):
        cell_fill = header_fill # 默认使用蓝色填充
        current_sas_name = None # 当前单元格对应的SAS列名

        # 如果单元格值为"更新情况（标记）"，则直接使用此名称
        if cell.value == '更新情况（标记）':
            current_sas_name = '更新情况（标记）' 
        # 如果当前列索引在sas_file_names列表范围内，获取对应的SAS列名
        elif col_idx - 1 < len(sas_file_names): 
            current_sas_name = sas_file_names[col_idx - 1] 
        
        # 不对"更新情况（标记）"列的表头进行列增删的染色
        if current_sas_name and current_sas_name != '更新情况（标记）': 
            # 如果当前SAS列名在删除列列表中，则使用缺失颜色填充
            if current_sas_name in del_sas_names:
                cell_fill = config.missing_sheet_tab_fill
            # 如果当前SAS列名在新增列列表中，则使用新增颜色填充
            elif current_sas_name in add_sas_names:
                cell_fill = config.new_sheet_tab_fill

        cell.fill = cell_fill # 应用计算出的填充色到表头单元格

    # 2. Sheet标签颜色已在`process_edc_multithreaded`中根据SheetProcessResult的`change_type`处理，此处无需再次处理。
    
    # 3. 数据行高亮 (Data row highlighting)
    mark_col_idx = None # "更新情况（标记）"列的Excel索引
    col_name_to_idx = {} # 列名到Excel索引的映射字典

    # 使用sas_file_names构建列名到索引的映射，确保与diff_info中的列名类型一致
    for col_idx, col_name in enumerate(sas_file_names, 1):
        if col_name == '更新情况（标记）':
            mark_col_idx = col_idx # 找到标记列的索引
        col_name_to_idx[col_name] = col_idx # 使用SASFieldName作为键
    
    if mark_col_idx is None: # 如果没有找到"更新情况（标记）"列，则无法进行后续高亮，直接返回
        return
    
    # 从第二行开始迭代所有数据行（因为第一行是表头）
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # 获取"更新情况（标记）"列中的单元格及其值
        mark_cell = row[mark_col_idx - 1] # openpyxl的row对象是0-based索引，所以需要-1
        mark_value = str(mark_cell.value).strip() if mark_cell.value is not None else ''      
        
        # 根据标记值应用不同的高亮样式
        if mark_value in ('新增', '删除'):
            mark_cell.font = Font(bold=True, color='FFFF0000') # 将标记列字体加粗并设为红色
            for cell in row: # 对于新增或删除的行，高亮显示整行的所有单元格
                cell.fill = config.highlight_fill     
        elif mark_value == '更新':
            mark_cell.font = Font(bold=True, color='FFFF0000') # 将标记列字体加粗并设为红色
            data_row_idx = int(row_idx - 2) # df_row_idx是DataFrame的0-based索引，Excel的row_idx是1-based
            if diff_info:
                # 兼容np.int64和int类型的key
                diff_keys = {int(k): v for k, v in diff_info.items()}
                if data_row_idx in diff_keys:
                    changed_cols = diff_keys[data_row_idx]
                    # print(f"高亮更新行: data_row_idx={data_row_idx}, changed_cols={list(changed_cols.keys())}, col_name_to_idx={col_name_to_idx}")
                    for col_name in changed_cols:
                        if col_name in col_name_to_idx:
                            col_excel_idx = col_name_to_idx[col_name] # 获取列的Excel索引（1-based）
                            try:
                                # cell = ws.cell(row=row_idx, column=col_excel_idx) # 获取对应的单元格
                                cell = row[col_excel_idx - 1] # 获取对应的单元格
                                cell.fill = config.highlight_fill # 应用高亮填充
                            except Exception as e:
                                if log_func:
                                    log_func(f"高亮单元格失败: row={row_idx}, col={col_name}, err={e}")
                                pass # 忽略单元格访问错误，继续处理 