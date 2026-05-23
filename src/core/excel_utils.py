import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet # 导入 Worksheet
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from ..utils.config_manager import ConfigManager # 修正导入路径

def replace_worksheet_headers(worksheet: Worksheet, new_headers: list) -> Worksheet:
    """
    替换 Openpyxl Worksheet 的表头。
    主要用于在Openpyxl工作表中设置新的列标题。
    
    参数:
        worksheet: 要处理的 Openpyxl Worksheet 对象。
        new_headers: 新表头列表（长度需与列数匹配）。
    
    返回:
        修改后的 Worksheet 对象。
    """
    # 验证新表头长度与列数匹配
    if len(new_headers) != worksheet.max_column:
        # 如果表头数量与工作表列数不匹配，则尝试调整新表头长度
        # 这可能发生在 dataframe_to_rows 使用后，或者新表头太短的情况下。
        if worksheet.max_column == 0: # 空白Sheet的情况
            worksheet.append(new_headers) # 直接添加新表头
            return worksheet

        print(f"警告: 新表头数量({len(new_headers)})与工作表列数({worksheet.max_column})不匹配。将尝试用新表头覆盖现有表头。")
        # 调整 new_headers 以匹配现有列数，如果需要则用空字符串填充
        if len(new_headers) < worksheet.max_column:
            new_headers.extend([''] * (worksheet.max_column - len(new_headers))) # 填充空字符串
        else: # 如果新表头更长，则截断
            new_headers = new_headers[:worksheet.max_column]

    # 处理合并单元格：保存合并信息并取消合并涉及表头行的单元格
    merged_ranges = []
    if worksheet.merged_cells.ranges:
        for merged_range in worksheet.merged_cells.ranges.copy():
            if merged_range.min_row == 1:  # 只处理涉及第一行（表头）的合并单元格
                merged_ranges.append(str(merged_range))
                worksheet.unmerge_cells(str(merged_range)) # 取消合并
    
    # 替换表头：遍历新表头列表，更新第一行的单元格值
    for col_idx, header in enumerate(new_headers, 1):
        cell = worksheet.cell(row=1, column=col_idx) # 获取第一行的单元格
        cell.value = header # 设置新的表头值
    
    # 重新应用合并单元格：恢复之前保存的表头合并信息
    for merged_range in merged_ranges:
        worksheet.merge_cells(merged_range) # 重新合并单元格
    
    return worksheet

def apply_highlight_to_worksheet(ws, config, sheet_type=None, diff_info=None, add_sas_names=[], del_sas_names=[], sas_file_names=[], log_func=None):
    """根据比对结果为工作表（包括表头、单元格和Sheet标签）应用高亮和颜色标记。
    "更新"的行只高亮有变化的单元格，"新增"或"删除"的行则高亮整行，并相应标记列的增删。
    
    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Openpyxl工作表对象，比对结果将应用到此工作表。
        config (ConfigManager): 配置管理器对象，包含颜色等设置。
        sheet_type (str, optional): Sheet的变更类型（'new', 'missing', 'data_changed'）。
        diff_info (dict, optional): 单元格级别的差异信息字典，格式为 {df_row_idx: {col_name: flag}}。
                                    这里的df_row_idx是DataFrame的0-based索引。
        add_sas_names (list, optional): 新增的SAS列名列表。
        del_sas_names (list, optional): 删除的SAS列名列表。
        sas_file_names (list, optional): 最终结果DataFrame的SAS列名列表（即表头行显示的列名）。
        log_func (callable, optional): 日志函数，用于记录操作信息。
    """

    # 1. 表头染色 (Header highlighting)
    # 默认表头填充色（蓝色），用于无特定增删标记的表头单元格
    header_fill = PatternFill(start_color='A6C9EC', end_color='A6C9EC', fill_type='solid')
    # 获取工作表的第一行（表头行），如果工作表为空则返回空列表
    first_row = ws[1] if ws.max_row > 0 else []

    # 遍历第一行的每个单元格（即表头单元格）
    for col_idx, cell in enumerate(first_row,1):
        cell_fill = header_fill # 默认使用蓝色填充
        current_sas_name = None # 当前单元格对应的SAS列名

        # 如果单元格值为"更新情况（标记）"，则直接使用此名称
        if cell.value == '更新情况（标记）':
            current_sas_name = '更新情况（标记）' 
        # 如果当前列索引在sas_file_names列表范围内，获取对应的SAS列名
        elif col_idx - 1 < len(sas_file_names): 
            current_sas_name = sas_file_names[col_idx - 1] 
        
        # 不对"更新情况（标记）"列的表头进行列增删的染色
        if current_sas_name and current_sas_name != '更新情况（标记）': 
            # 如果当前SAS列名在删除列列表中，则使用缺失颜色填充
            if current_sas_name in del_sas_names:
                cell_fill = config.missing_sheet_tab_fill
            # 如果当前SAS列名在新增列列表中，则使用新增颜色填充
            elif current_sas_name in add_sas_names:
                cell_fill = config.new_sheet_tab_fill

        cell.fill = cell_fill # 应用计算出的填充色到表头单元格

    # 2. Sheet标签颜色已在`process_edc_multithreaded`中根据SheetProcessResult的`change_type`处理，此处无需再次处理。
    
    # 3. 数据行高亮 (Data row highlighting)
    mark_col_idx = None # "更新情况（标记）"列的Excel索引
    col_name_to_idx = {} # 列名到Excel索引的映射字典

    # 使用sas_file_names构建列名到索引的映射，确保与diff_info中的列名类型一致
    for col_idx, col_name in enumerate(sas_file_names, 1):
        if col_name == '更新情况（标记）':
            mark_col_idx = col_idx # 找到标记列的索引
        col_name_to_idx[col_name] = col_idx # 使用SASFieldName作为键
    
    if mark_col_idx is None: # 如果没有找到"更新情况（标记）"列，则无法进行后续高亮，直接返回
        return
    
    # 从第二行开始迭代所有数据行（因为第一行是表头）
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        # 获取"更新情况（标记）"列中的单元格及其值
        mark_cell = row[mark_col_idx - 1] # openpyxl的row对象是0-based索引，所以需要-1
        mark_value = str(mark_cell.value).strip() if mark_cell.value is not None else ''      
        
        # 根据标记值应用不同的高亮样式
        if mark_value in ('新增', '删除'):
            mark_cell.font = Font(bold=True, color='FFFF0000') # 将标记列字体加粗并设为红色
            for cell in row: # 对于新增或删除的行，高亮显示整行的所有单元格
                cell.fill = config.highlight_fill     
        elif mark_value == '更新':
            mark_cell.font = Font(bold=True, color='FFFF0000') # 将标记列字体加粗并设为红色
            data_row_idx = int(row_idx - 2) # df_row_idx是DataFrame的0-based索引，而Excel的row_idx是1-based
            if diff_info:
                # 兼容np.int64和int类型的key
                # diff_keys = {int(k): v for k, v in diff_info.items()} # 引入numpy依赖，暂时注释
                diff_keys = {int(k): v for k, v in diff_info.items()} # 替换为没有numpy依赖的版本
                if data_row_idx in diff_keys:
                    changed_cols = diff_keys[data_row_idx]
                    # print(f"高亮更新行: data_row_idx={data_row_idx}, changed_cols={list(changed_cols.keys())}, col_name_to_idx={col_name_to_idx}")
                    for col_name in changed_cols:
                        if col_name in col_name_to_idx:
                            col_excel_idx = col_name_to_idx[col_name] # 获取列的Excel索引（1-based）
                            try:
                                # cell = ws.cell(row=row_idx, column=col_excel_idx) # 获取对应的单元格
                                cell = row[col_excel_idx - 1] # 获取对应的单元格
                                cell.fill = config.highlight_fill # 应用高亮填充
                            except Exception as e:
                                if log_func:
                                    log_func(f"高亮单元格失败: row={row_idx}, col={col_name}, err={e}")
                                pass # 忽略单元格访问错误，继续处理 