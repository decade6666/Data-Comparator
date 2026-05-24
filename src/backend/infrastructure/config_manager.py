import os
from typing import Dict, List, Optional

from openpyxl.styles import PatternFill

from ...shared.contracts import ParameterColors, ParameterDocument


class ConfigManager:
    def __init__(self) -> None:
        self.anchor_row_num = 1
        self.header_row_num = 1
        self.anchor_row_content = "SASFieldName"
        self.header_row_content = "SASFieldLabel"
        self.exclude_sheets: List[str] = []
        self.common_cols_to_drop: List[str] = []
        self.default_keys: List[str] = []
        self.sheet_key_map: Dict[str, List[str]] = {}
        self.highlight_fill: Optional[PatternFill] = None
        self.missing_sheet_tab_fill: Optional[PatternFill] = None
        self.new_sheet_tab_fill: Optional[PatternFill] = None
        cores = os.cpu_count() or 1
        self.max_workers = max(1, cores - 1)
        self.merge_deleted_data = True

    def _to_argb(self, color: str) -> str:
        text = color.replace("#", "")
        if len(text) == 6:
            return "FF" + text.upper()
        if len(text) == 8:
            return text.upper()
        return "FFFFFFFF"

    def update_from_parameters(
        self, parameters: ParameterDocument, colors: ParameterColors
    ) -> None:
        self.anchor_row_num = parameters.get("anchor_row_num", 1)
        self.header_row_num = parameters.get("header_row_num", 1)
        self.anchor_row_content = parameters.get("anchor_row_content", "SASFieldName")
        self.header_row_content = parameters.get("header_row_content", "SASFieldLabel")
        self.exclude_sheets = list(parameters.get("exclude_sheets", []))
        self.common_cols_to_drop = list(parameters.get("common_cols", []))
        self.default_keys = list(parameters.get("default_keys", []))
        self.sheet_key_map = dict(parameters.get("sheet_key_map", {}))
        self.max_workers = parameters.get(
            "max_workers", max(1, (os.cpu_count() or 1) - 1)
        )
        self.merge_deleted_data = parameters.get("merge_deleted_data", True)

        highlight_color = colors.get("highlight_fill", "#FFE5E5")
        missing_color = colors.get("missing_sheet_tab", "#DC143C")
        new_color = colors.get("new_sheet_tab", "#00FF00")

        self.highlight_fill = PatternFill(
            start_color=self._to_argb(highlight_color),
            end_color=self._to_argb(highlight_color),
            fill_type="solid",
        )
        self.missing_sheet_tab_fill = PatternFill(
            start_color=self._to_argb(missing_color),
            end_color=self._to_argb(missing_color),
            fill_type="solid",
        )
        self.new_sheet_tab_fill = PatternFill(
            start_color=self._to_argb(new_color),
            end_color=self._to_argb(new_color),
            fill_type="solid",
        )
