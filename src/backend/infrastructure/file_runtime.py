import datetime
import os
import shutil
from typing import Any, List, Optional, Tuple

import appdirs
import pandas as pd
from openpyxl import load_workbook

from ...shared.contracts import LogFunc
from ...shared.log_utils import log
from ..domain.processing_control import check_stop_frequently
from .xlsx_filter_cleaner import (
    FilterCleanupResult,
    NotAnOoxmlPackageError,
    remove_filters,
)


def _safe_log(log_func: Optional[LogFunc], message: str) -> None:
    if not log_func:
        return
    try:
        log_func(message)
    except Exception:
        pass


def check_and_remove_file_protection(
    file_path: str,
    exclude_sheets: List[str],
    log_func: Optional[LogFunc],
    stop_flag: Optional[Any] = None,
    work_dir: Optional[str] = None,
) -> Tuple[bool, bool, str, Optional[FilterCleanupResult]]:
    """Copy an input file and clean supported OOXML filter state on the copy.

    ``exclude_sheets`` remains in the signature for compatibility with existing
    callers. Sheet selection is performed later by the comparison domain layer.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")

    base, ext = os.path.splitext(file_path)
    temp_app_dir = work_dir or get_app_temp_dir()
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S%f")
    original_filename = os.path.basename(base)
    new_file_name = f"{original_filename}_nofilter_{timestamp}{ext}"
    new_file_path = os.path.join(temp_app_dir, new_file_name)

    try:
        os.makedirs(temp_app_dir, exist_ok=True)
        shutil.copy2(file_path, new_file_path)
        check_stop_frequently(log_func, stop_flag)
    except InterruptedError:
        _cleanup_temp_copy(new_file_path, log_func)
        raise
    except Exception as exc:
        _cleanup_temp_copy(new_file_path, log_func)
        _safe_log(log_func, f"❌ 创建副本失败: {str(exc)}")
        raise

    cleanup_result: Optional[FilterCleanupResult] = None
    try:
        cleanup_result = remove_filters(
            new_file_path,
            log_func=log_func,
            stop_flag=stop_flag,
        )
    except InterruptedError:
        _cleanup_temp_copy(new_file_path, log_func)
        raise
    except NotAnOoxmlPackageError as exc:
        _safe_log(log_func, f"ℹ️ 非 OOXML 包（.xls 等），跳过筛选器清理: {str(exc)}")
    except Exception:
        _cleanup_temp_copy(new_file_path, log_func)
        raise
    else:
        if cleanup_result.rewritten:
            _safe_log(
                log_func,
                "✅ 已清除筛选器：工作表 "
                f"{cleanup_result.sheet_autofilters_removed} 处、表格 "
                f"{cleanup_result.table_autofilters_removed} 处，恢复隐藏行 "
                f"{cleanup_result.hidden_rows_restored} 行",
            )
        else:
            _safe_log(
                log_func,
                "ℹ️ 未发现自动筛选器，跳过重写 "
                f"（扫描 {cleanup_result.parts_scanned} 个部件）",
            )

    return False, False, new_file_path, cleanup_result


def _cleanup_temp_copy(path: str, log_func: Optional[LogFunc]) -> None:
    try:
        if os.path.exists(path):
            os.remove(path)
    except OSError as exc:
        _safe_log(log_func, f"⚠️ 中断后清理临时副本失败: {str(exc)}")


def validate_excel_file(
    file_path: str,
    log_func,
) -> Tuple[bool, Optional[str]]:
    try:
        if not os.path.exists(file_path):
            error = f"文件不存在: {file_path}"
            log(error, log_func)
            return False, error

        file_size = os.path.getsize(file_path)
        if file_size == 0:
            error = f"文件为空: {file_path}"
            log(error, log_func)
            return False, error

        engines = ["openpyxl", "xlrd"]
        validation_errors: List[str] = []

        for engine in engines:
            try:
                if engine == "xlrd" and file_path.endswith(".xlsx"):
                    continue
                pd.read_excel(file_path, sheet_name=0, nrows=1, engine=engine)
                return True, None
            except Exception as engine_error:
                error_msg = str(engine_error)
                validation_errors.append(f"{engine}: {error_msg}")

                if (
                    "invalid XML" in error_msg
                    or "could not read worksheets" in error_msg
                ):
                    log("⚠️ 检测到XML格式错误，建议修复文件后重试", log_func)
                continue

        combined_error = "; ".join(validation_errors)
        log("⚠️ 所有验证引擎都失败，但文件可能仍可修复", log_func)
        return False, combined_error
    except Exception as e:
        error = f"文件验证过程出错: {str(e)}"
        log(error, log_func)
        return False, error


def get_sheet_names(file_path: str, log_func) -> List[str]:
    try:
        wb = load_workbook(file_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        return sheet_names
    except Exception as e:
        log_func(f"⚠️ 无法获取文件 {os.path.basename(file_path)} 的Sheet名称: {str(e)}")
        try:
            excel_file = pd.ExcelFile(file_path, engine="openpyxl")
            sheet_names = excel_file.sheetnames
            excel_file.close()
            return sheet_names
        except Exception as e_pd:
            log_func(
                f"❌ 无法获取文件 {os.path.basename(file_path)} 的Sheet名称"
                f"（pandas回退也失败）: {str(e_pd)}"
            )
            return []


def get_app_data_dir() -> str:
    """应用持久数据根目录（数据库、按用户隔离的数据）。"""
    override = os.environ.get("DATASET_COMPARATOR_DATA_DIR", "").strip()
    if override:
        os.makedirs(override, exist_ok=True)
        return override
    appname = "PyDataCompare"
    appauthor = "YourCompanyOrAuthor"
    data_dir = appdirs.user_data_dir(appname, appauthor)
    os.makedirs(data_dir, exist_ok=True)
    return data_dir


def get_user_data_dir(user_id: int) -> str:
    """某用户的隔离数据根目录：<data>/users/<uid>。"""
    user_dir = os.path.join(get_app_data_dir(), "users", str(user_id))
    os.makedirs(user_dir, exist_ok=True)
    return user_dir


def get_app_temp_dir() -> str:
    appname = "PyDataCompare"
    appauthor = "YourCompanyOrAuthor"
    temp_dir = appdirs.user_data_dir(appname, appauthor)
    temp_sub_dir = os.path.join(temp_dir, "temp")
    os.makedirs(temp_sub_dir, exist_ok=True)
    return temp_sub_dir


def cleanup_nofilter_files(
    log_func=None,
    work_dir: Optional[str] = None,
) -> int:
    """清理指定工作目录下的临时副本；缺省时清理全局临时目录。

    work_dir 用于按任务隔离清理，避免并发任务互删中间文件。
    """
    removed_count = 0
    try:
        temp_dir = work_dir or get_app_temp_dir()
        if not os.path.isdir(temp_dir):
            return 0
        for name in os.listdir(temp_dir):
            if "_nofilter_" in name and name.lower().endswith(
                (".xlsx", ".xlsm", ".xls")
            ):
                fpath = os.path.join(temp_dir, name)
                try:
                    if os.path.isfile(fpath):
                        os.remove(fpath)
                        removed_count += 1
                except Exception as e:
                    if log_func:
                        log_func(f"⚠️ 删除临时缓存文件失败: {fpath}，原因: {e}")
        if log_func:
            log_func(f"🧹 已清理 nofilter 缓存文件 {removed_count} 个")
    except Exception as e:
        if log_func:
            log_func(f"⚠️ 清理临时缓存文件时出错: {e}")
    return removed_count
