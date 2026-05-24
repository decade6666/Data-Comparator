from datetime import date, datetime
from typing import List, Optional

import pandas as pd
from openpyxl import load_workbook

from .processing_control import check_stop_frequently


def read_single_sheet_from_excel(
    file_path: str,
    sheet_name: str,
    anchor_row_num: int,
    header_row_num: int,
    log_func,
    cols_to_drop: Optional[List[str]] = None,
):
    """
    在一个独立的线程中，读取Excel文件中指定Sheet的数据，并处理SASFieldLabel和SASFieldName。
    此函数会在读取完成后关闭文件，确保资源及时释放。
    Args:
        file_path (str): Excel文件路径。
        sheet_name (str): 需要读取的Sheet名称。
        anchor_row_num (int): 锚点所在行的行号（1-based）。
        header_row_num (int): 生成文件表头所在行的行号（1-based）。
        log_func (callable): 日志函数。
        cols_to_drop (list, optional): 在读取后需要删除的列名列表。
    Returns:
        pd.DataFrame: 读取到的数据DataFrame，包含SAS元数据作为attrs。
            如果读取失败或Sheet不存在，则返回None。
    """
    try:
        # 使用openpyxl的load_workbook以read_only模式打开，data_only=True确保读取的是显示值而非公式
        wb = load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            wb.close()
            return None  # Sheet不存在，返回None

        ws = wb[sheet_name]  # 获取指定Sheet

        # 获取SASFieldLabel和SASFieldName（列标签和列名）
        sas_field_label = []
        sas_field_name = []

        # 直接使用传入的行号
        label_row_idx = header_row_num
        name_row_idx = anchor_row_num
        # 读取到表头和锚点行中的最大行数
        header_rows_to_read = max(label_row_idx, name_row_idx)

        def _normalize_value(value):
            # Keep numeric types, normalize dates to string, otherwise stringify.
            if value is None:
                return None
            if isinstance(value, (datetime, date)):
                return str(value)
            if isinstance(value, (int, float)) and not isinstance(value, bool):
                return value
            try:
                return str(value)
            except Exception:
                return None

        header_values_list = []
        # 使用 iter_rows 仅读取需要的头部行，避免一次性读取过多数据
        for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=header_rows_to_read, values_only=True)
        ):  # 从第1行开始读取到指定表头行
            if r_idx % 50 == 0:
                check_stop_frequently(log_func)
            header_values_list.append(
                [_normalize_value(value) for value in row]
            )  # 提取单元格值

        # 根据索引获取SASFieldLabel和SASFieldName
        if label_row_idx > 0 and label_row_idx <= len(header_values_list):
            sas_field_label = [
                str(val) if val is not None else ""
                for val in header_values_list[label_row_idx - 1]
            ]

        if name_row_idx > 0 and name_row_idx <= len(header_values_list):
            sas_field_name = [
                str(val) if val is not None else ""
                for val in header_values_list[name_row_idx - 1]
            ]
            # 检测锚点行是否存在重复列名
            non_empty_names = [
                n.strip() for n in sas_field_name if isinstance(n, str) and n.strip()
            ]
            if len(non_empty_names) != len(set(non_empty_names)):
                # 统计重复项
                from collections import Counter

                cnt = Counter(non_empty_names)
                dup_list = [k for k, v in cnt.items() if v > 1]
                preview = ", ".join(dup_list[:10])
                msg = (
                    f"Sheet [{sheet_name}] 的锚点行(第 {name_row_idx} 行)存在重复内容: {preview}"
                    + (" 等" if len(dup_list) > 10 else "")
                )
                log_func(f"❌ {msg}")
                raise ValueError(msg)

        # 逐行读取数据，直到遇到第一个空行，或者读取到最大行数（防止超大文件耗尽内存）
        data_rows = []
        max_rows_to_read = 1000000  # 最大读取行数限制

        # 从数据开始行读取（表头行之后）
        for row_idx_absolute, row in enumerate(
            ws.iter_rows(min_row=header_rows_to_read + 1, values_only=True),
            start=header_rows_to_read + 1,
        ):
            if (row_idx_absolute - (header_rows_to_read + 1)) % 100 == 0:
                check_stop_frequently(log_func)
            row_values = [
                _normalize_value(value) for value in row
            ]  # 获取当前行的所有单元格值

            # 检查是否是空行（所有值都为None或空字符串），如果是则停止读取
            if all(
                v is None or (isinstance(v, str) and not v.strip()) for v in row_values
            ):
                log_func(
                    f"检测到Sheet [{sheet_name}] 在第 {row_idx_absolute} 行遇到空行，停止读取。"
                )
                break

            data_rows.append(row_values)  # 添加数据行

            if len(data_rows) > max_rows_to_read:  # 如果读取行数超过限制，强制停止
                log_func(
                    f"⚠️ Sheet [{sheet_name}] 读取行数超过 {max_rows_to_read}，强制停止读取。"
                )
                break

        wb.close()  # 读取完指定Sheet后立即关闭文件，释放文件句柄和内存

        # 处理SASFieldName和SASFieldLabel的兼容性问题
        # 确保sas_field_name和sas_field_label长度一致，或用索引补齐
        if (
            not sas_field_name
        ):  # 如果没有定义SASFieldName行，则使用SASFieldLabel作为列名
            sas_field_name = sas_field_label

        # 如果SASFieldName仍然为空（例如，非常规表头或空Sheet），则使用默认列名
        if not sas_field_name:
            # 如果有数据行，尝试从数据行推断列数
            if data_rows:
                num_cols = max(len(row) for row in data_rows)  # 获取最长行的列数
                sas_field_name = [
                    f"Unnamed_{i}" for i in range(num_cols)
                ]  # 生成Unnamed列名
            else:
                sas_field_name = []  # 仍然为空

        if not sas_field_label:  # 如果没有定义SASFieldLabel，则使用SASFieldName作为标签
            sas_field_label = sas_field_name

        # 确保列名长度与实际数据列数一致，进行填充或截断
        if data_rows:
            actual_num_cols = max(len(row) for row in data_rows)  # 实际数据列数
            if len(sas_field_name) < actual_num_cols:
                sas_field_name.extend(
                    [
                        f"Unnamed_{i}"
                        for i in range(len(sas_field_name), actual_num_cols)
                    ]
                )  # 补充Unnamed列名
            if len(sas_field_label) < actual_num_cols:
                sas_field_label.extend(
                    [
                        f"Unnamed_{i}"
                        for i in range(len(sas_field_label), actual_num_cols)
                    ]
                )  # 补充空标签

        # 创建DataFrame，指定列名
        # 修正：即使data_rows为空，也使用sas_field_name作为columns，以保留表头信息
        df = pd.DataFrame(data_rows, columns=sas_field_name)

        # 提前删除指定列
        if cols_to_drop:
            # log_func(f"   [{sheet_name}] 正在删除指定列: {cols_to_drop}")
            original_columns_before_drop = list(df.columns)  # 记录删除前的列名
            df = df.drop(columns=cols_to_drop, axis=1, errors="ignore")

            # 更新 sas_field_name 和 sas_field_label，只保留未删除的列
            new_sas_field_name = []
            new_sas_field_label = []
            name_to_label_map = {
                name: label for name, label in zip(sas_field_name, sas_field_label)
            }

            for col in df.columns:  # 遍历删除后的当前列名
                new_sas_field_name.append(col)
                new_sas_field_label.append(
                    name_to_label_map.get(col, col)
                )  # 从原始映射中获取标签

            sas_field_name = new_sas_field_name
            sas_field_label = new_sas_field_label

        # 将SASFieldLabel和SASFieldName作为DataFrame的属性存储（Pandas的.attrs属性）
        df.attrs["sas_file_name"] = df.columns.tolist()  # 存储最终的列名
        df.attrs["sas_file_label"] = sas_field_label[
            : len(df.columns)
        ]  # 存储对应的列标签

        # 创建SAS列名到列标签的映射字典，基于存储在attrs中的信息
        sas_name_to_label = {}
        for i, name in enumerate(df.attrs["sas_file_name"]):
            if i < len(df.attrs["sas_file_label"]):
                sas_name_to_label[name] = df.attrs["sas_file_label"][i]
            else:
                sas_name_to_label[name] = name  # 如果没有对应的标签，使用列名本身

        sas_name_to_label["更新情况（标记）"] = (
            "更新情况（标记）"  # 确保"更新情况（标记）"列的映射存在
        )
        df.attrs["sas_name_to_label"] = sas_name_to_label  # 存储映射字典

        # 初始化用于后续比对的额外列信息（新增/删除列）
        if not hasattr(df, "attrs"):
            df.attrs = {}
        df.attrs["_add_sas_names"] = []  # 初始化新增列列表
        df.attrs["_del_sas_names"] = []  # 初始化删除列列表

        # 返回DataFrame
        return df

    except InterruptedError:
        try:
            if "wb" in locals() and wb:
                wb.close()
        except Exception:
            pass
        raise
    except Exception as e:
        log_func(f"❌ 读取Sheet [{sheet_name}] 失败: {str(e)}")  # 记录错误日志
        try:
            if "wb" in locals() and wb:  # 尝试关闭可能已打开的工作簿
                wb.close()
        except Exception:
            pass  # 忽略关闭失败的错误
        return None
