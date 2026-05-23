import os
import sys
import shutil
import zipfile
import tempfile
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Protection, Font, PatternFill, Border, Alignment, Side, NamedStyle, Color, GradientFill
from openpyxl.drawing.image import Image
from .logger import log
from ..core.excel_header_utils import read_single_sheet_from_excel

import re # Add regular expression module for text processing
import json # Add json module for handling json files
import unicodedata # For handling Unicode characters
import appdirs # Import the appdirs library
import datetime # Add datetime module for timestamp

try:
    import win32com.client as win32
    import pythoncom
except Exception:
    win32 = None
    pythoncom = None

def get_resource_path(relative_path):
    """获取资源文件的绝对路径，支持开发环境和PyInstaller打包后的环境。
    Args:
        relative_path (str): 相对于程序根目录的资源文件路径。
    Returns:
        str: 资源文件的绝对路径。
    """
    try:
        # PyInstaller会将资源文件打包到临时文件夹，并通过sys._MEIPASS提供路径
        base_path = sys._MEIPASS
    except Exception:
        # 如果不是打包环境（例如在开发环境中运行），则使用当前文件的目录作为基准路径
        current_dir = os.path.dirname(os.path.abspath(__file__))
        base_path = os.path.join(current_dir, os.pardir) # 返回到src目录
    
    return os.path.join(base_path, relative_path) # 拼接绝对路径

# 设置窗口图标
def set_window_icon(window):
    """为Tkinter窗口设置应用程序图标。"""
    try:
        icon_files = [
            'app_icon.ico',
            os.path.join('assets', 'icons', 'app_icon.ico'),
        ] # 定义可能的图标文件名列表（兼容开发与打包路径）
        
        icon_loaded = False
        for icon_file in icon_files:
            icon_path = get_resource_path(icon_file) # 获取图标的绝对路径
            if os.path.exists(icon_path): # 检查图标文件是否存在
                window.iconbitmap(icon_path) # 设置窗口图标
                print(f"[OK] 成功加载图标: {icon_file}")
                icon_loaded = True
                break # 找到并加载成功后退出循环
        
        if not icon_loaded:
            print("[WARN] 未找到图标文件，使用默认图标") # 如果所有图标文件都未找到
            
    except Exception as e:
        print(f"[ERROR] 设置图标失败: {str(e)}") # 打印设置图标失败的错误信息

# 将'更新情况（标记）'列移动到第1列
def reorder_columns_with_update_mark_first(df):
    """将DataFrame中的'更新情况（标记）'列移动到第一列。"""
    try:
        if '更新情况（标记）' in df.columns: # 检查是否存在该列
            cols = df.columns.tolist() # 获取所有列名列表
            cols.remove('更新情况（标记）') # 从原位置移除
            cols.insert(0, '更新情况（标记）') # 插入到列表的第一个位置
            df = df[cols] # 根据新的列顺序重新选择列，实现列的移动
        return df
    except Exception as e:
        print(f"⚠️ 调整列顺序时出错: {str(e)}") # 打印警告信息
        return df # 发生错误时返回原始DataFrame

# 检查文件是否被标记为受保护（例如来自Internet的文件），如有则尝试解除保护。
# 如果需要解除保护，则先另存为新文件（加_nofilter后缀），并返回新路径。
def check_and_remove_file_protection(file_path, exclude_sheets, log_func):
    """
    处理Excel文件：
    1. 另存为新文件（`_nofilter`后缀），以避免修改原始文件。
    2. 尝试解除文件保护（仅适用于Windows NTFS文件系统，通过删除Zone.Identifier ADS）。
    3. 清除文件中的自动筛选器。
    
    Args:
        file_path (str): 原始Excel文件路径。
        exclude_sheets (list): 需要排除的工作表名称列表。
        log_func (callable): 日志函数，用于输出详细处理信息。
        
    Returns:
        tuple: (was_protected, removed_successfully, final_path, actually_removed)
            - was_protected (bool): 文件在处理前是否被标记为受保护。
            - removed_successfully (bool): 文件保护是否成功移除。
            - final_path (str): 处理后的新文件路径。
            - actually_removed (list): 兼容字段，现不再删除工作表，始终为空列表。
            
    注意：文件保护解除功能仅适用于Windows NTFS文件系统。
    """   
    # 验证输入文件是否存在
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"文件不存在: {file_path}")
    
    # ======== 步骤1: 另存文件 ========
    base, ext = os.path.splitext(file_path) # 分离文件名和扩展名
    # new_file_path = base + '_nofilter' + ext # 构建新文件路径
    
    # 将_nofilter文件保存到AppData中的临时目录
    temp_app_dir = get_app_temp_dir()
    # 生成一个基于原始文件名和时间戳的唯一文件名，避免冲突
    timestamp = datetime.datetime.now().strftime("%Y%m%d%H%M%S%f")
    original_filename = os.path.basename(base)
    new_file_name = f"{original_filename}_nofilter_{timestamp}{ext}"
    new_file_path = os.path.join(temp_app_dir, new_file_name)
    
    try:
        shutil.copy2(file_path, new_file_path) # 复制文件，保留元数据
        # 复制可能较慢，复制后立刻检查一次停止
        try:
            check_stop_frequently(log_func)
        except Exception:
            # 保持原异常向上抛，由调用方捕获
            raise
    except Exception as e:
        log_func(f"❌ 创建副本失败: {str(e)}")
        raise # 复制失败则抛出异常
    
    # ======== 步骤2: 解除文件保护 ========
    zone_id_stream = f"{file_path}:Zone.Identifier" # Windows文件流，用于标记文件来源
    is_protected = False # 初始假定文件未受保护
    protection_removed = False # 初始假定保护未被移除
    
    try:
        # 检查是否存在Zone.Identifier流，并读取其内容判断是否受保护（ZoneId=3表示来自Internet）
        if os.path.exists(zone_id_stream):
            with open(zone_id_stream, 'r') as f:
                content = f.read()
            is_protected = "[ZoneTransfer]" in content and "ZoneId=3" in content
        else:
            is_protected = False # 流不存在则认为未受保护
    except Exception:
        is_protected = True  # 无法确定（例如权限问题），保守假定受保护
    
    if is_protected:
        try:
            os.remove(zone_id_stream) # 尝试删除Zone.Identifier流以解除保护
            protection_removed = True # 标记保护已成功移除
        except Exception as e:
            protection_removed = False # 移除失败
            log_func(f"⚠️ 解除文件保护失败: {str(e)}") # 记录警告
    else:
        log_func("ℹ️ 文件未受保护，继续处理...") # 文件不受保护，正常处理
    
    actually_removed = [] # 兼容字段，现在不删除工作表，始终为空列表

    # 使用 pywin32 清除自动筛选器
    try:
        # 先使用 pywin32 清除所有工作表的自动筛选器
        if win32 is None:
            raise ImportError("pywin32 未安装或不可用")
        # 初始化 COM（在多线程环境下尤其重要）
        try:
            pythoncom.CoInitialize()
        except Exception:
            pass

        excel_app = None
        wb_com = None
        filters_cleared_by_pywin32 = False
        created_standalone_app = False
        try:
            dispatch_method = getattr(win32, "DispatchEx", None)
            if dispatch_method is not None:
                excel_app = dispatch_method("Excel.Application")
                created_standalone_app = True
            else:
                excel_app = win32.Dispatch("Excel.Application")
            excel_app.Visible = False
            excel_app.DisplayAlerts = False

            check_stop_frequently(log_func)

            wb_com = excel_app.Workbooks.Open(new_file_path, UpdateLinks=0, ReadOnly=False)
            for ws_com in wb_com.Worksheets:
                try:
                    check_stop_frequently(log_func)
                except Exception:
                    if wb_com is not None:
                        wb_com.Close(SaveChanges=False)
                    if excel_app is not None and created_standalone_app:
                        excel_app.Quit()
                    raise

                # 若存在已应用的筛选，先显示全部数据
                try:
                    if getattr(ws_com, "FilterMode", False):
                        try:
                            ws_com.ShowAllData()
                        except Exception:
                            try:
                                ws_com.AutoFilter.ShowAllData()
                            except Exception:
                                pass
                        filters_cleared_by_pywin32 = True
                except Exception:
                    pass

                # 关闭工作表级的筛选器（移除筛选箭头）
                try:
                    if getattr(ws_com, "AutoFilterMode", False):
                        ws_com.AutoFilterMode = False
                        filters_cleared_by_pywin32 = True
                except Exception:
                    pass

                # 针对表(ListObject)的筛选器也一并关闭
                try:
                    list_objects = getattr(ws_com, "ListObjects", None)
                    if list_objects is not None:
                        for lo in list_objects:
                            try:
                                if getattr(lo, "ShowAutoFilter", None) is not None and lo.ShowAutoFilter:
                                    lo.ShowAutoFilter = False
                                    filters_cleared_by_pywin32 = True
                                try:
                                    lo.AutoFilter.ShowAllData()
                                except Exception:
                                    pass
                            except Exception:
                                pass
                except Exception:
                    pass

            wb_com.Save()
        finally:
            try:
                if wb_com is not None:
                    wb_com.Close(SaveChanges=False)
            except Exception:
                pass
            try:
                if excel_app is not None and created_standalone_app:
                    excel_app.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass

        if filters_cleared_by_pywin32:
            log_func("✅ 成功通过pywin32清除自动筛选器。")
        else:
            log_func("ℹ️ 文件中未发现自动筛选器，跳过清除。")


    except Exception as e:
        log_func(f"⚠️ 主要预处理（pywin32清除筛选器或删除Sheet）失败: {str(e)}，尝试回退方法...")
        # Fallback path: If pywin32 operations failed
        # 1. Try to clear filters using the original (slower) method
        try:
            remove_auto_filters_from_xlsx(new_file_path, new_file_path, log_func)
            log_func("✅ 成功通过备用方法清除自动筛选器。")
        except Exception as fallback_e:
            log_func(f"❌ 备用筛选器清除失败，跳过继续处理: {str(fallback_e)}")

    return is_protected, protection_removed, new_file_path, actually_removed

# 验证Excel文件的完整性和格式
def validate_excel_file(file_path, log_func):
    """
    验证Excel文件的完整性和格式。尝试使用不同的引擎读取文件，以判断其可读性。
    Args:
        file_path (str): 需要验证的Excel文件路径。
        log_func (callable): 日志函数，用于输出详细信息。
    Returns:
        tuple: (is_valid, error_details)
            - is_valid (bool): 文件是否有效且可读。
            - error_details (str): 如果无效，返回具体的错误信息。
    """

    try:
        # 检查文件是否存在
        if not os.path.exists(file_path):
            error = f"文件不存在: {file_path}"
            log(error, log_func)
            return False, error
        
        # 检查文件大小是否为0
        file_size = os.path.getsize(file_path)
        if file_size == 0:
            error = f"文件为空: {file_path}"
            log(error, log_func)
            return False, error
        
        # 尝试使用不同的方法读取文件（openpyxl 和 xlrd）
        engines = ['openpyxl', 'xlrd'] # 定义尝试的引擎列表
        validation_errors = [] # 存储验证过程中遇到的错误
        
        for engine in engines:
            try:
                if engine == 'xlrd' and file_path.endswith('.xlsx'):
                    continue  # xlrd不支持.xlsx格式，跳过
                # 尝试读取文件的第一个Sheet的第一行，快速验证文件结构
                test_df = pd.read_excel(file_path, sheet_name=0, nrows=1, engine=engine)
                return True, None # 成功读取，文件有效
            except Exception as engine_error:
                error_msg = str(engine_error)
                validation_errors.append(f"{engine}: {error_msg}") # 记录引擎特定的错误
                
                # 检查是否是XML格式错误，给出修复建议
                if "invalid XML" in error_msg or "could not read worksheets" in error_msg:
                    log(f"⚠️ 检测到XML格式错误，建议修复文件后重试", log_func)
                continue # 继续尝试下一个引擎
        
        # 如果所有引擎都失败
        combined_error = "; ".join(validation_errors) # 合并所有错误信息
        log(f"⚠️ 所有验证引擎都失败，但文件可能仍可修复", log_func)
        return False, combined_error # 返回验证失败和合并后的错误信息
        
    except Exception as e:
        error = f"文件验证过程出错: {str(e)}"
        log(error, log_func)
        return False, error # 返回验证失败和内部错误信息

# 全局停止标志，用于快速检查
_global_stop_flag = None

def set_global_stop_flag(stop_flag):
    """设置全局停止标志，用于在多线程操作中控制程序的停止。"""
    global _global_stop_flag
    _global_stop_flag = stop_flag # 将传入的Event对象赋值给全局变量

def check_stop_frequently(log_func, stop_flag=None):
    """高频率停止检查，用于关键处理路径，确保程序能及时响应停止指令。"""
    global _global_stop_flag
    current_flag = stop_flag or _global_stop_flag # 优先使用传入的stop_flag，否则使用全局的
    if current_flag and current_flag.is_set(): # 检查停止标志是否被设置
        log("处理已被用户停止", log_func) # 记录日志
        raise InterruptedError("用户停止了操作") # 抛出中断异常，终止当前操作

# 检查是否应该停止处理 - 带计数器优化
def check_stop(log_func, stop_flag=None, check_counter=None):
    """检查是否应该停止处理，带计数器优化性能，减少不必要的频繁检查。
    在耗时循环中，可以每隔N次操作检查一次停止标志。
    Args:
        log_func (callable): 日志函数。
        stop_flag (threading.Event, optional): 停止标志Event对象。
        check_counter (list, optional): 包含一个整数的列表，用于计数检查频率。
    """
    global _global_stop_flag
    current_flag = stop_flag or _global_stop_flag # 优先使用传入的stop_flag，否则使用全局的
    
    # 使用计数器减少检查频率，提升性能
    if check_counter is not None:
        if check_counter[0] % 100 != 0:  # 每100次操作检查一次（可调整频率）
            check_counter[0] += 1
            return
        check_counter[0] += 1
    
    if current_flag and current_flag.is_set(): # 检查停止标志是否被设置
        log("处理已被用户停止", log_func) # 记录日志
        raise InterruptedError("用户停止了操作") # 抛出中断异常

#更新进度条 - 线程安全版本
def update_progress(msg, progress=None, progress_func=None):
    """线程安全的进度更新函数。
    通过回调函数将进度信息传递给GUI，避免直接在工作线程中操作GUI。
    Args:
        msg (str): 进度消息文本。
        progress (int, optional): 进度百分比 (0-100)。
        progress_func (callable, optional): 实际执行GUI进度更新的回调函数。
    """
    if progress_func: # 如果提供了进度更新函数
        try:
            progress_func(msg, progress) # 调用回调函数更新GUI
        except Exception as e:
            # 如果进度更新失败，不中断主流程，只是打印错误到控制台
            print(f"进度更新异常: {str(e)}")


# 在不加载完整数据的情况下获取Excel文件的所有Sheet名称
def get_sheet_names(file_path: str, log_func):
    """
    仅读取Excel文件的所有Sheet名称，不加载Sheet的全部数据，以提高效率。
    尝试使用openpyxl，如果失败则回退到pandas。
    Args:
        file_path (str): Excel文件路径。
        log_func (callable): 日志函数。
    Returns:
        list: Sheet名称列表。如果获取失败，返回空列表。
    """
    try:
        wb = load_workbook(file_path, read_only=True) # 以只读模式加载工作簿
        sheet_names = wb.sheetnames # 获取所有Sheet名称
        wb.close() # 及时关闭工作簿，释放资源
        return sheet_names
    except Exception as e:
        log_func(f"⚠️ 无法获取文件 {os.path.basename(file_path)} 的Sheet名称: {str(e)}")
        # 如果无法通过openpyxl获取，尝试使用pandas（但pandas可能会加载更多数据）
        try:
            excel_file = pd.ExcelFile(file_path, engine='openpyxl') # 使用pandas的ExcelFile对象
            sheet_names = excel_file.sheetnames
            excel_file.close() # 关闭ExcelFile对象
            return sheet_names
        except Exception as e_pd:
            log_func(f"❌ 无法获取文件 {os.path.basename(file_path)} 的Sheet名称（pandas回退也失败）: {str(e_pd)}")
            return [] # 彻底失败，返回空列表


def get_app_temp_dir():
    """
    获取应用程序的临时文件目录，位于用户AppData（或等效）目录下。
    """
    appname = "PyDataCompare" # 定义应用程序名称
    appauthor = "YourCompanyOrAuthor" # 定义应用程序作者（可选，但建议提供）
    temp_dir = appdirs.user_data_dir(appname, appauthor)
    temp_sub_dir = os.path.join(temp_dir, "temp")
    os.makedirs(temp_sub_dir, exist_ok=True)
    return temp_sub_dir

def cleanup_nofilter_files(log_func=None):
    """删除应用临时目录中缓存的所有 *_nofilter_* Excel 副本文件。
    返回删除的文件数量。
    当未传入 log_func 时静默执行，不写日志。
    """
    removed_count = 0
    try:
        temp_dir = get_app_temp_dir()
        if not os.path.isdir(temp_dir):
            return 0
        for name in os.listdir(temp_dir):
            # 仅清理由本程序生成的 _nofilter_ 副本文件
            if '_nofilter_' in name and name.lower().endswith(('.xlsx', '.xlsm')):
                fpath = os.path.join(temp_dir, name)
                try:
                    if os.path.isfile(fpath):
                        os.remove(fpath)
                        removed_count += 1
                except Exception as e:
                    if log_func:
                        log_func(f"⚠️ 删除临时缓存文件失败: {fpath}，原因: {e}")
        if log_func:
            log_func(f"🧹 已清理 nofilter 缓存文件 {removed_count} 个")
    except Exception as e:
        if log_func:
            log_func(f"⚠️ 清理临时缓存文件时出错: {e}")
    return removed_count

# Traverse worksheet XML files, delete <autoFilter> tags
def remove_auto_filters_from_xlsx(file_path, output_path=None,log_message=None):
    """
    遍历Excel文件（.xlsx）中的工作表XML文件，删除<autoFilter>标签。
    此函数用于清除Excel文件中的自动筛选器，以便程序能正常读取和处理数据。
    
    Args:
        file_path (str): 输入的Excel文件路径。
        output_path (str, optional): 输出的Excel文件路径。如果为None，则默认为file_path。
        log_message (callable, optional): 日志函数，用于记录操作信息。
    
    注意：此操作会解压Excel文件，修改XML，然后重新压缩。
    """
    # 确定输出路径，如果未指定则默认为输入文件路径
    output_path = output_path or file_path.replace(".xlsx", ".xlsx")

    # 使用临时目录解压Excel文件（Excel文件本质是ZIP压缩包）
    # 将使用系统的AppData文件夹作为临时目录
    temp_app_dir = get_app_temp_dir()
    unique_temp_id = os.urandom(8).hex() # 生成一个唯一的ID以避免文件名冲突
    tmpdirname = os.path.join(temp_app_dir, f"excel_extract_{unique_temp_id}")
    os.makedirs(tmpdirname, exist_ok=True) # 确保临时目录存在
    
    try:
        with zipfile.ZipFile(file_path, 'r') as zip_ref:
            zip_ref.extractall(tmpdirname) # 解压所有文件到临时目录
        # 解压完成后立即检查一次停止
        check_stop_frequently(log_message)

        # 定位到包含工作表XML文件的目录
        sheet_dir = os.path.join(tmpdirname, "xl", "worksheets")
        for filename in os.listdir(sheet_dir):
            # 遍历每个XML前检查一次停止
            check_stop_frequently(log_message)
            if filename.startswith("sheet") and filename.endswith(".xml"):
                sheet_path = os.path.join(sheet_dir, filename)
                tree = ET.parse(sheet_path) # 解析XML文件
                root = tree.getroot() # 获取XML根元素

                # 定义Excel XML中使用的默认命名空间
                ns = {'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'}

                # 查找并删除<autoFilter>标签
                auto_filter = root.find("main:autoFilter", ns)
                if auto_filter is not None:
                    root.remove(auto_filter) # 移除autoFilter元素
                    tree.write(sheet_path, encoding="utf-8", xml_declaration=True) # 将修改后的XML写回文件

        # 重新压缩临时目录中的文件，生成新的Excel文件
        with zipfile.ZipFile(output_path, 'w', zipfile.ZIP_DEFLATED) as new_zip:
            for foldername, subfolders, filenames in os.walk(tmpdirname):
                # 每处理一个目录层检查一次停止
                check_stop_frequently(log_message)
                for filename in filenames:
                    # 每处理一定数量文件检查一次停止
                    check_stop_frequently(log_message)
                    file_path_inner = os.path.join(foldername, filename)
                    arcname = os.path.relpath(file_path_inner, tmpdirname) # 获取文件在ZIP中的相对路径
                    new_zip.write(file_path_inner, arcname) # 添加文件到新的ZIP包中
    finally:
        # 清理生成的临时目录
        if os.path.exists(tmpdirname):
            shutil.rmtree(tmpdirname) 