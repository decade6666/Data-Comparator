from openpyxl.styles import PatternFill
import os

class ConfigManager:
    """配置管理器类，用于集中管理和更新应用程序的各种配置参数。
    这些参数包括锚点行号、表头行号、需要排除的Sheet、通用删除列、默认锚点、特定Sheet的锚点映射以及颜色设置。
    """
    def __init__(self):
        # 默认配置初始化
        self.anchor_row_num = 1 # 锚点所在行号
        self.header_row_num = 1 # 生成文件表头所在行号
        self.anchor_row_content = 'SASFieldName'
        self.header_row_content = 'SASFieldLabel'
        self.exclude_sheets = [] # 需要排除的Sheet列表
        self.common_cols_to_drop = [] # 需要删除的通用列列表
        self.default_keys = [] # 默认的锚点列列表
        self.sheet_key_map = {} # 特定Sheet的锚点映射字典
        self.highlight_fill = None # 高亮填充样式
        self.missing_sheet_tab_fill = None # 缺失Sheet标签填充样式
        self.new_sheet_tab_fill = None # 新增Sheet标签填充样式
        cores = os.cpu_count() or 1
        self.max_workers = max(1, cores - 1)
        self.merge_deleted_data = True
    
    def _to_argb(self, color):
        """将#RRGGBB或RRGGBB颜色字符串转换为8位ARGB格式（如FFFFE5E5）"""
        c = color.replace('#', '')
        if len(c) == 6:
            return 'FF' + c.upper()
        elif len(c) == 8:
            return c.upper()
        else:
            return 'FFFFFFFF'  # fallback to white
    
    def update_from_parameters(self, parameters, colors):
        """从参数对象和颜色字典更新配置。"""
        self.anchor_row_num = parameters.get('anchor_row_num', 1)
        self.header_row_num = parameters.get('header_row_num', 1)
        self.anchor_row_content = parameters.get('anchor_row_content', 'SASFieldName')
        self.header_row_content = parameters.get('header_row_content', 'SASFieldLabel')
        self.exclude_sheets = parameters.get('exclude_sheets', []) # 获取排除的Sheet列表
        self.common_cols_to_drop = parameters.get('common_cols', []) # 获取通用删除列列表
        self.default_keys = parameters.get('default_keys', []) # 获取默认锚点列表
        self.sheet_key_map = parameters.get('sheet_key_map', {}) # 获取特定Sheet的锚点映射
        self.max_workers = parameters.get('max_workers', max(1, (os.cpu_count() or 1) - 1))
        self.merge_deleted_data = parameters.get('merge_deleted_data', True)
        
        # 更新颜色填充样式，使用PatternFill对象，确保为8位ARGB
        highlight_color = colors.get('highlight_fill', '#FFE5E5') # 获取高亮颜色，默认#FFE5E5
        missing_color = colors.get('missing_sheet_tab', '#DC143C') # 获取缺失颜色，默认#DC143C
        new_color = colors.get('new_sheet_tab', '#00FF00') # 获取新增颜色，默认#00FF00
        
        self.highlight_fill = PatternFill(
            start_color=self._to_argb(highlight_color),
            end_color=self._to_argb(highlight_color),
            fill_type='solid'
        )
        
        self.missing_sheet_tab_fill = PatternFill(
            start_color=self._to_argb(missing_color),
            end_color=self._to_argb(missing_color),
            fill_type='solid'
        )
        
        self.new_sheet_tab_fill = PatternFill(
            start_color=self._to_argb(new_color),
            end_color=self._to_argb(new_color),
            fill_type='solid'
        ) 